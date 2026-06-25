"""Excel report generation from registered modeling artifacts."""

from __future__ import annotations

import json
import time
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd

from risk_model_workbench.config import load_yaml


REPORT_SHEETS = [
    "模型描述",
    "重要变量",
    "Top变量WOE",
    "变量筛选过程和模型参数",
    "模型效果-每月效果",
    "模型效果-模型sloping",
    "模型效果-意愿交叉风险（DEV-OOS）",
    "模型稳定性",
]
GCARD_SUMMARY_SHEET = "Summary"

SCORE_COLUMNS = ["model_score"]
VERSION_LABELS = {
    "model_score": "本轮模型",
}
REPORT_TITLE = "Model Report"
PROJECT_DISPLAY_NAME = "Model"
SEGMENT_FILES = {
    "全客群": "all",
    "老户次新": "e2e3",
    "流失户": "b2",
}


def generate_excel_report(
    *,
    eval_dir: str | Path,
    train_dir: str | Path,
    input_dir: str | Path,
    feature_dir: str | Path,
    output_path: str | Path,
    project_dir: str | Path | None = None,
    report_config: dict[str, Any] | None = None,
) -> Path:
    """Generate a formatted Excel model report from standard artifact folders."""
    from openpyxl import Workbook

    eval_dir = Path(eval_dir)
    train_dir = Path(train_dir)
    input_dir = Path(input_dir)
    feature_dir = Path(feature_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    context = _build_report_context(project_dir=project_dir, eval_dir=eval_dir, report_config=report_config)
    global SCORE_COLUMNS, VERSION_LABELS, REPORT_TITLE, PROJECT_DISPLAY_NAME
    previous_context = (SCORE_COLUMNS, VERSION_LABELS, REPORT_TITLE, PROJECT_DISPLAY_NAME)
    SCORE_COLUMNS = context["score_columns"]
    VERSION_LABELS = context["score_labels"]
    REPORT_TITLE = context["report_title"]
    PROJECT_DISPLAY_NAME = context["project_display_name"]

    try:
        run_dir = eval_dir.parent if eval_dir.name == "evaluation" else output_path.parent.parent
        sample_dir = run_dir / "sample_check"
        include_gcard_summary = bool(context.get("include_gcard_summary"))

        wb = Workbook()
        ws = wb.active
        if include_gcard_summary:
            ws.title = GCARD_SUMMARY_SHEET
            for sheet_name in REPORT_SHEETS:
                wb.create_sheet(sheet_name)
        else:
            ws.title = REPORT_SHEETS[0]
            for sheet_name in REPORT_SHEETS[1:]:
                wb.create_sheet(sheet_name)

        if include_gcard_summary:
            _build_gcard_summary_sheet(
                wb[GCARD_SUMMARY_SHEET],
                train_dir=train_dir,
                eval_dir=eval_dir,
                feature_dir=feature_dir,
            )

        _build_description_sheet(
            wb["模型描述"],
            train_dir=train_dir,
            eval_dir=eval_dir,
            input_dir=input_dir,
            sample_dir=sample_dir,
            feature_dir=feature_dir,
        )
        _build_features_sheet(wb["重要变量"], train_dir=train_dir, feature_dir=feature_dir)
        _build_woe_sheet(wb["Top变量WOE"], train_dir=train_dir, report_dir=output_path.parent)
        _build_screening_params_sheet(wb["变量筛选过程和模型参数"], train_dir=train_dir, feature_dir=feature_dir)
        _build_monthly_effect_sheet(wb["模型效果-每月效果"], eval_dir=eval_dir)
        _build_sloping_sheet(wb["模型效果-模型sloping"], eval_dir=eval_dir)
        _build_intent_risk_sheet(wb["模型效果-意愿交叉风险（DEV-OOS）"], eval_dir=eval_dir)
        _build_stability_sheet(wb["模型稳定性"], eval_dir=eval_dir)

        for worksheet in wb.worksheets:
            _finalize_sheet(worksheet)

        wb.save(str(output_path))
        _write_missing_results_doc(output_path, train_dir=train_dir, eval_dir=eval_dir)
        _write_model_reports(
            output_path=output_path,
            train_dir=train_dir,
            eval_dir=eval_dir,
            feature_dir=feature_dir,
            sample_dir=sample_dir,
            include_gcard_summary=include_gcard_summary,
        )
    finally:
        SCORE_COLUMNS, VERSION_LABELS, REPORT_TITLE, PROJECT_DISPLAY_NAME = previous_context
    return output_path


def _build_report_context(
    *,
    project_dir: str | Path | None,
    eval_dir: Path,
    report_config: dict[str, Any] | None,
) -> dict[str, Any]:
    project_path = Path(project_dir) if project_dir else _infer_project_dir(eval_dir)
    project_config = _load_first_yaml(project_path, ["project.yml", "project.yaml"]) if project_path else {}
    evaluate_config = _load_first_yaml(project_path / "configs", ["evaluate.yaml", "evaluate.yml"]) if project_path else {}
    loaded_report_config = report_config or (_load_first_yaml(project_path / "configs", ["report.yaml", "report.yml"]) if project_path else {})

    project_display_name = project_config.get("project", {}).get("display_name") or (project_path.name if project_path else "Model")
    eval_cfg = evaluate_config.get("evaluation", {}) if isinstance(evaluate_config.get("evaluation"), dict) else {}
    score_columns = eval_cfg.get("score_columns") or ["model_score"]
    score_labels = {"model_score": "本轮模型"}
    configured_labels = eval_cfg.get("score_labels") or {}
    if isinstance(configured_labels, dict):
        score_labels.update({str(key): str(value) for key, value in configured_labels.items()})
    for score_column in score_columns:
        score_labels.setdefault(str(score_column), str(score_column))

    report_root = loaded_report_config.get("report", {}) if isinstance(loaded_report_config.get("report"), dict) else {}
    report_title = report_root.get("title") or f"{project_display_name}模型报告"
    project_root = project_config.get("project", {}) if isinstance(project_config.get("project"), dict) else {}
    project_name = str(project_root.get("name") or (project_path.name if project_path else ""))
    project_template = str(project_root.get("template") or "")
    include_gcard_summary = (
        "gcard_v6" in [str(column) for column in score_columns]
        and (
            project_name == "2026-05-fujie-gcard-v1"
            or project_template == "fujie-gcard"
            or str(project_display_name) == "复借G卡"
        )
    )
    return {
        "project_display_name": str(project_display_name),
        "score_columns": [str(column) for column in score_columns],
        "score_labels": score_labels,
        "report_title": str(report_title),
        "include_gcard_summary": include_gcard_summary,
    }


def _infer_project_dir(eval_dir: Path) -> Path | None:
    resolved = eval_dir.resolve()
    for parent in [resolved, *resolved.parents]:
        if parent.name == "runs":
            return parent.parent
    return None


def _load_first_yaml(directory: Path, names: list[str]) -> dict[str, Any]:
    for name in names:
        path = directory / name
        if path.exists():
            try:
                return load_yaml(path)
            except (OSError, ValueError):
                return {}
    return {}


def _build_gcard_summary_sheet(
    ws,
    *,
    train_dir: Path,
    eval_dir: Path,
    feature_dir: Path,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    compare_score = "gcard_v6"
    compare_label = VERSION_LABELS.get(compare_score, compare_score)
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    overall = _read_csv(eval_dir / "overall_metrics.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    importance = _read_csv(train_dir / "feature_importance.csv")
    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    score_distribution = _read_csv(eval_dir / "score_bin_distribution_by_month_by_version.csv")

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=13)
    title_cell = ws.cell(2, 2)
    title_cell.value = "复借G卡模型对比 Summary"
    title_cell.font = Font(name="楷体", size=18, bold=True, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, text in [
        (3, f"口径：新版模型（model_score） vs 旧版全客群模型（{compare_label}）；分客群仅作为切片效果，不含分客群专属模型。"),
        (4, "更新日期：" + time.strftime("%Y-%m-%d") + "；指标来源：当前 run 已注册 train/evaluation artifacts。"),
    ]:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        cell = ws.cell(row, 2)
        cell.value = text
        cell.font = Font(name="楷体", size=10, italic=True, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 6
    row = _write_summary_heading(ws, row, "一、模型描述")
    _write_table(
        ws,
        row,
        "模型与标签口径",
        _gcard_model_description_frame(train_dir=train_dir, feature_dir=feature_dir, run_config=run_config, metrics=metrics),
        start_col=1,
        apply_color_scale=False,
    )
    row = _write_table(
        ws,
        row,
        "样本切分",
        _gcard_sample_split_frame(overall=overall, monthly=monthly, compare_label=compare_label),
        start_col=8,
        apply_color_scale=False,
    )

    row = _write_summary_heading(ws, row, "二、重要变量")
    row = _write_note(ws, row, "本轮仅训练单一全客群主模型；Top变量WOE 详见“Top变量WOE”sheet。")
    row = _write_table(ws, row, "Top 20 重要变量", _gcard_top_features_frame(importance), apply_color_scale=False)

    row = _write_summary_heading(ws, row, "三、变量筛选过程和模型参数")
    row = _write_summary_table_pair(
        ws,
        row,
        left_title="变量筛选过程",
        left_frame=_gcard_screening_summary_frame(train_dir=train_dir, feature_dir=feature_dir),
        right_title="模型参数",
        right_frame=_gcard_model_params_frame(run_config=run_config, metrics=metrics),
    )
    row = _write_note(ws, row, "连续性说明：本页 Summary 只读取当前 run 已登记产物；缺失链路不补造指标。")

    row = _write_summary_heading(ws, row, "四、模型效果")
    row = _write_note(ws, row, f"以下只展示新版模型 vs 旧版全客群 {compare_label}；AUC 和 KS 按手工版样式拆成横向两个小表。")
    row = _write_summary_table_pair(
        ws,
        row,
        left_title="整体效果对比 AUC",
        left_frame=_gcard_metric_comparison_frame(overall, row_col="final_flag", metric="auc", row_label="样本", compare_score=compare_score),
        right_title="整体效果对比 KS",
        right_frame=_gcard_metric_comparison_frame(overall, row_col="final_flag", metric="ks", row_label="样本", compare_score=compare_score),
    )
    if segment is not None and not segment.empty:
        row = _write_note(ws, row, "分客群结果是效果切片，不代表本轮训练了老户/次新/流失户专属模型。")
        oot_oos_segment = _ordered_subset(segment, "segment", ["全客群", "老户次新", "老户", "次新", "流失户"])
        oot_oos_segment = oot_oos_segment[oot_oos_segment["final_flag"] == "OOT-OOS"] if "final_flag" in oot_oos_segment.columns else oot_oos_segment
        row = _write_summary_table_pair(
            ws,
            row,
            left_title="OOT-OOS 分客群切片 AUC",
            left_frame=_gcard_metric_comparison_frame(oot_oos_segment, row_col="segment", metric="auc", row_label="客群", compare_score=compare_score, include_bad_rate=True),
            right_title="OOT-OOS 分客群切片 KS",
            right_frame=_gcard_metric_comparison_frame(oot_oos_segment, row_col="segment", metric="ks", row_label="客群", compare_score=compare_score, include_bad_rate=True),
        )
    if monthly is not None and not monthly.empty:
        oos_monthly = monthly[monthly["final_flag"].isin(["DEV-OOS", "OOT-OOS"])].copy() if "final_flag" in monthly.columns else monthly.copy()
        oos_monthly["_period_label"] = oos_monthly["final_flag"].astype(str) + " " + oos_monthly["mdl_month"].astype(str)
        row = _write_summary_table_pair(
            ws,
            row,
            left_title="OOS 按月 AUC",
            left_frame=_gcard_metric_comparison_frame(oos_monthly, row_col="_period_label", metric="auc", row_label="样本月份", compare_score=compare_score, include_bad_rate=True),
            right_title="OOS 按月 KS",
            right_frame=_gcard_metric_comparison_frame(oos_monthly, row_col="_period_label", metric="ks", row_label="样本月份", compare_score=compare_score, include_bad_rate=True),
        )

    row = _write_summary_heading(ws, row, "模型sloping（OOT-OOS，10箱完整表现）")
    row = _write_gcard_sloping_summary(ws, row, eval_dir=eval_dir, compare_score=compare_score)

    intent_frame = _gcard_intent_summary_frame(eval_dir=eval_dir, compare_score=compare_score)
    if not intent_frame.empty:
        row = _write_summary_heading(ws, row, "意愿 x 资产评级风险矩阵（DEV-OOS，已补齐 by segment / version / zc）")
        row = _write_table(ws, row, "全客群 intent=意愿档，zc=合计", intent_frame, apply_color_scale=False)

    row = _write_summary_heading(ws, row, "五、模型稳定性")
    row = _write_note(ws, row, "PSI 同步展示月度汇总和按月/版本/decile 分箱明细覆盖情况；PSI 保留三位小数。")
    row = _write_table(ws, row, "月度 PSI", _gcard_psi_summary_frame(psi=psi, compare_score=compare_score), apply_color_scale=False)
    row = _write_table(ws, row, "稳定性分箱明细覆盖", _gcard_stability_coverage_frame(score_distribution=score_distribution, compare_score=compare_score), apply_color_scale=False)

    row = _write_table(
        ws,
        row,
        "补充数据覆盖核对",
        _gcard_coverage_frame(eval_dir=eval_dir, feature_dir=feature_dir),
        apply_color_scale=False,
    )
    _write_note(ws, row, "证据提示：本页不补造缺失结果；当前 run audit 仍应以 rmw run audit 输出为准。")
    ws.freeze_panes = "A6"
    for column, width in {
        "A": 16,
        "B": 20,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 4,
        "H": 16,
        "I": 20,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
    }.items():
        ws.column_dimensions[column].width = width


def _write_summary_heading(ws, row: int, title: str) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    cell = ws.cell(row, 1)
    cell.value = title
    cell.font = Font(name="楷体", size=13, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 2


def _write_summary_table_pair(
    ws,
    row: int,
    *,
    left_title: str,
    left_frame: pd.DataFrame,
    right_title: str,
    right_frame: pd.DataFrame,
) -> int:
    left_end = _write_table(ws, row, left_title, left_frame, start_col=1, apply_color_scale=False)
    right_end = _write_table(ws, row, right_title, right_frame, start_col=8, apply_color_scale=False)
    return max(left_end, right_end)


def _gcard_model_description_frame(*, train_dir: Path, feature_dir: Path, run_config: dict[str, Any], metrics: dict[str, Any]) -> pd.DataFrame:
    warnings = _feature_list_warnings(train_dir)
    rows = [
        ("标签字段", run_config.get("label_column", "ftr_30d_ord_flag")),
        ("Y定义", "观察日后30天内是否发起订单"),
        ("训练/验证/OOS", f"{_fmt_list(run_config.get('train_values', []))} / {_fmt_list(run_config.get('valid_values', []))} / {_fmt_list(run_config.get('oos_values', []))}"),
        ("算法", run_config.get("algorithm", "N/A")),
        ("入模特征数", _feature_count(train_dir, feature_dir) or run_config.get("actual_feature_count", "N/A")),
        ("对比对象", "本轮 model_score vs 旧版全客群 G卡V6"),
        ("Best iteration", run_config.get("best_iteration", "N/A")),
        ("Valid AUC / KS", f"{_fmt_metric(metrics.get('valid_auc'))} / {_fmt_metric(metrics.get('valid_ks'))}"),
    ]
    if warnings:
        rows.append(("风险提示", warnings[0]))
    return pd.DataFrame(rows, columns=["项目", "内容"])


def _gcard_sample_split_frame(*, overall: pd.DataFrame | None, monthly: pd.DataFrame | None, compare_label: str) -> pd.DataFrame:
    if overall is None or overall.empty:
        return pd.DataFrame()
    rows = []
    ordered = _ordered_subset(overall, "final_flag", ["DEV", "DEV-OOS", "OOT", "OOT-OOS"])
    for row in ordered.itertuples(index=False):
        values = row._asdict()
        final_flag = values.get("final_flag")
        rows.append(
            {
                "样本": final_flag,
                "样本数": values.get("n_samples"),
                "30天发起量": values.get("positive"),
                "30天发起率": values.get("bad_rate"),
                "月份范围": _gcard_month_range(monthly, str(final_flag)),
                "基准": compare_label,
            }
        )
    return pd.DataFrame(rows)


def _gcard_month_range(monthly: pd.DataFrame | None, final_flag: str) -> str:
    if monthly is None or monthly.empty or "final_flag" not in monthly.columns or "mdl_month" not in monthly.columns:
        return "N/A"
    subset = monthly[monthly["final_flag"] == final_flag].copy()
    if subset.empty:
        return "N/A"
    months = sorted(str(value) for value in subset["mdl_month"].dropna().unique().tolist())
    if not months:
        return "N/A"
    return months[0] if len(months) == 1 else f"{months[0]} ~ {months[-1]}"


def _gcard_top_features_frame(importance: pd.DataFrame | None) -> pd.DataFrame:
    if importance is None or importance.empty:
        return pd.DataFrame()
    top = importance.head(20).copy().reset_index(drop=True)
    total_gain = pd.to_numeric(importance.get("gain"), errors="coerce").fillna(0).sum()
    gain = pd.to_numeric(top.get("gain"), errors="coerce").fillna(0)
    top["index"] = range(1, len(top) + 1)
    top["varname"] = top["feature"]
    top["desc"] = top["feature"]
    top["gain占比"] = gain / total_gain if total_gain else 0
    top["累计占比"] = top["gain占比"].cumsum()
    return top[[col for col in ["index", "varname", "desc", "gain", "gain占比", "累计占比", "split"] if col in top.columns]]


def _gcard_screening_summary_frame(*, train_dir: Path, feature_dir: Path) -> pd.DataFrame:
    process = _read_gcard_feature_screening_process(feature_dir)
    if process.get("screening_rows"):
        previous: int | None = None
        rows = []
        for idx, item in enumerate(process["screening_rows"]):
            remaining = item.get("remaining_features")
            try:
                remaining_int = int(remaining)
            except (TypeError, ValueError):
                remaining_int = None
            if idx == 0:
                step = "初始"
                dropped = ""
            else:
                step = f"D{idx:02d}"
                dropped = previous - remaining_int if previous is not None and remaining_int is not None else ""
            rows.append(
                {
                    "步骤": step,
                    "筛选方法": item.get("method"),
                    "剩余变量数": remaining,
                    "本步剔除": dropped,
                    "来源": item.get("source"),
                }
            )
            previous = remaining_int
        return pd.DataFrame(rows)
    stage_summary = _read_json(feature_dir / "stage_summary.json") or _read_json(feature_dir / "feature_stage_summary.json")
    if stage_summary:
        return _screening_steps_frame(stage_summary, feature_dir).rename(columns={"剩余变量个数": "剩余变量数"})
    return _training_feature_preparation_frame(train_dir)


def _read_gcard_feature_screening_process(feature_dir: Path) -> dict[str, Any]:
    for path in [
        feature_dir / "feature_screening_process.json",
        feature_dir / "feature_screening_process_v2.json",
        feature_dir.parent.parent.parent / "reports" / "feature_screening_process.json",
    ]:
        if path.exists():
            return _read_json(path)
    return {}


def _gcard_model_params_frame(*, run_config: dict[str, Any], metrics: dict[str, Any]) -> pd.DataFrame:
    params = run_config.get("params", {}) if isinstance(run_config.get("params"), dict) else {}
    rows = [{"参数": key, "取值": value} for key, value in list(params.items())[:8]]
    rows.extend(
        [
            {"参数": "训练耗时", "取值": f"{metrics.get('train_time_seconds', 'N/A')}s"},
            {"参数": "Train/Valid AUC gap", "取值": metrics.get("auc_gap", "N/A")},
        ]
    )
    return pd.DataFrame(rows)


def _gcard_metric_comparison_frame(
    frame: pd.DataFrame | None,
    *,
    row_col: str,
    metric: str,
    row_label: str,
    compare_score: str,
    include_bad_rate: bool = False,
) -> pd.DataFrame:
    if frame is None or frame.empty or row_col not in frame.columns:
        return pd.DataFrame()
    metric_label = metric.upper()
    model_col = f"model_score_{metric}"
    compare_col = f"{compare_score}_{metric}"
    if model_col not in frame.columns or compare_col not in frame.columns:
        return pd.DataFrame()
    display = frame.copy()
    columns: dict[str, Any] = {
        row_label: display[row_col],
    }
    if "n_samples" in display.columns:
        columns["样本数"] = display["n_samples"]
    if include_bad_rate and "bad_rate" in display.columns:
        columns["30天发起率"] = display["bad_rate"]
    columns[f"本轮{metric_label}"] = display[model_col]
    columns[f"{VERSION_LABELS.get(compare_score, compare_score)} {metric_label}"] = display[compare_col]
    columns[f"{metric_label}提升"] = pd.to_numeric(display[model_col], errors="coerce") - pd.to_numeric(display[compare_col], errors="coerce")
    return pd.DataFrame(columns)


def _ordered_subset(frame: pd.DataFrame, column: str, order: list[str]) -> pd.DataFrame:
    if frame.empty or column not in frame.columns:
        return frame.copy()
    ordered = frame.copy()
    ordered["_order"] = ordered[column].map({value: idx for idx, value in enumerate(order)}).fillna(len(order))
    return ordered.sort_values(["_order", column]).drop(columns="_order").reset_index(drop=True)


def _write_gcard_sloping_summary(ws, row: int, *, eval_dir: Path, compare_score: str) -> int:
    row = _write_note(ws, row, "2025-12 至 2026-01 30天发起 OOT-OOS")
    for segment_label, segment_key in SEGMENT_FILES.items():
        model_frame = _gcard_sloping_source_frame(eval_dir=eval_dir, segment_label=segment_label, segment_key=segment_key, score_column="model_score")
        compare_frame = _gcard_sloping_source_frame(eval_dir=eval_dir, segment_label=segment_label, segment_key=segment_key, score_column=compare_score)
        if model_frame.empty and compare_frame.empty:
            continue
        row = _write_note(ws, row, f"在{segment_label}效果")
        row = _write_summary_table_pair(
            ws,
            row,
            left_title=VERSION_LABELS.get("model_score", "本轮模型"),
            left_frame=model_frame,
            right_title=VERSION_LABELS.get(compare_score, compare_score),
            right_frame=compare_frame,
        )
    return row


def _gcard_sloping_source_frame(*, eval_dir: Path, segment_label: str, segment_key: str, score_column: str) -> pd.DataFrame:
    versioned = _read_csv(eval_dir / "decile_lift_bins_by_version.csv")
    if versioned is not None and not versioned.empty and {"segment", "final_flag", "score_version"}.issubset(versioned.columns):
        subset = versioned[
            (versioned["segment"] == segment_label)
            & (versioned["final_flag"] == "OOT-OOS")
            & (versioned["score_version"] == score_column)
        ].copy()
        if not subset.empty:
            return _sloping_display_frame(subset)
    path = eval_dir / f"decile_lift_{segment_key}_{score_column}.csv"
    if score_column == "model_score" and not path.exists():
        path = eval_dir / f"decile_lift_{segment_key}.csv"
    frame = _read_csv(path)
    return _sloping_display_frame(frame) if frame is not None and not frame.empty else pd.DataFrame()


def _gcard_intent_summary_frame(*, eval_dir: Path, compare_score: str) -> pd.DataFrame:
    distribution = _read_csv(eval_dir / "intent_zc_segment_distribution_by_version.csv")
    ftr_rate = _read_csv(eval_dir / "intent_zc_segment_ftr_rate_by_version.csv")
    amount_risk = _read_csv(eval_dir / "intent_zc_segment_amount_risk_by_version.csv")
    if distribution is None or ftr_rate is None or amount_risk is None:
        return pd.DataFrame()
    rows = []
    for intent in ["低意愿", "中意愿", "高意愿"]:
        row: dict[str, Any] = {"意愿档": intent}
        for score_column, prefix in [("model_score", "本轮"), (compare_score, VERSION_LABELS.get(compare_score, compare_score))]:
            row[f"{prefix}样本数"] = _intent_metric_value(distribution, intent, "n_samples", score_column)
            row[f"{prefix}金额逾期率"] = _intent_metric_value(amount_risk, intent, "amount_overdue_rate", score_column)
            row[f"{prefix}人头风险率"] = _intent_metric_value(amount_risk, intent, "head_risk_rate", score_column)
            row[f"{prefix}FTR"] = _intent_metric_value(ftr_rate, intent, "ftr_30d_rate", score_column)
        rows.append(row)
    return pd.DataFrame(rows)


def _intent_metric_value(frame: pd.DataFrame, intent: str, value_col: str, score_column: str) -> Any:
    required = {"segment", "final_flag", "score_version", "intent_level", "zc_level", value_col}
    if frame.empty or not required.issubset(frame.columns):
        return None
    subset = frame[
        (frame["segment"] == "全客群")
        & (frame["final_flag"] == "DEV-OOS")
        & (frame["score_version"] == score_column)
        & (frame["intent_level"] == intent)
        & (frame["zc_level"].astype(str) == "合计")
    ]
    if subset.empty:
        return None
    return subset.iloc[0][value_col]


def _gcard_psi_summary_frame(*, psi: pd.DataFrame | None, compare_score: str) -> pd.DataFrame:
    if psi is None or psi.empty:
        return pd.DataFrame()
    rows = []
    for month in sorted(str(value) for value in psi["month"].dropna().unique().tolist()):
        row = {"月份": month}
        for score_column, prefix in [("model_score", "本轮"), (compare_score, VERSION_LABELS.get(compare_score, compare_score))]:
            subset = psi[(psi["month"].astype(str) == month) & (psi["score_column"] == score_column)] if "score_column" in psi.columns else psi[psi["month"].astype(str) == month]
            if subset.empty:
                continue
            row[f"{prefix} PSI"] = subset.iloc[0].get("psi")
            row[f"{prefix}样本数"] = subset.iloc[0].get("n_samples")
        rows.append(row)
    return pd.DataFrame(rows)


def _gcard_stability_coverage_frame(*, score_distribution: pd.DataFrame | None, compare_score: str) -> pd.DataFrame:
    if score_distribution is None or score_distribution.empty or "score_column" not in score_distribution.columns:
        return pd.DataFrame(
            [
                {
                    "版本": "model_score",
                    "分箱行数": "待补",
                    "月份数": "待补",
                    "decile数": "待补",
                    "样本行合计": "待补",
                    "最大月PSI": "待补",
                    "最大PSI月份": "待补",
                },
                {
                    "版本": compare_score,
                    "分箱行数": "待补",
                    "月份数": "待补",
                    "decile数": "待补",
                    "样本行合计": "待补",
                    "最大月PSI": "待补",
                    "最大PSI月份": "待补",
                },
            ]
        )
    rows = []
    for score_column in ["model_score", compare_score]:
        subset = score_distribution[score_distribution["score_column"] == score_column].copy()
        if subset.empty:
            continue
        max_month = None
        max_psi = None
        if "month_psi" in subset.columns:
            max_row = subset.sort_values("month_psi", ascending=False).iloc[0]
            max_month = max_row.get("mdl_month")
            max_psi = max_row.get("month_psi")
        rows.append(
            {
                "版本": score_column,
                "分箱行数": len(subset),
                "月份数": subset["mdl_month"].nunique() if "mdl_month" in subset.columns else "N/A",
                "decile数": subset["score_decile"].nunique() if "score_decile" in subset.columns else "N/A",
                "样本行合计": pd.to_numeric(subset.get("n_samples"), errors="coerce").sum() if "n_samples" in subset.columns else "N/A",
                "最大月PSI": max_psi,
                "最大PSI月份": max_month,
            }
        )
    return pd.DataFrame(rows)


def _gcard_coverage_frame(*, eval_dir: Path, feature_dir: Path) -> pd.DataFrame:
    checks = [
        ("特征筛选全过程", any((feature_dir / name).exists() for name in ["feature_screening_process.json", "feature_screening_process_v2.json", "stage_summary.json", "feature_stage_summary.json"]), "feature_selection/*.json/txt 已落到当前 run。"),
        ("稳定性分箱明细", (eval_dir / "score_bin_distribution_by_month_by_version.csv").exists(), "score_bin_distribution_by_month_by_version.csv 覆盖按月、版本、decile。"),
        ("意愿风险矩阵", all((eval_dir / name).exists() for name in ["intent_zc_segment_distribution_by_version.csv", "intent_zc_segment_ftr_rate_by_version.csv", "intent_zc_segment_amount_risk_by_version.csv"]), "distribution/ftr/amount 三个矩阵覆盖分群、版本、意愿层、资产评级层。"),
        ("报告刷新范围", True, "本次刷新由 rmw report 统一生成；未直接重跑训练/评估。"),
    ]
    return pd.DataFrame(
        {
            "数据项": [name for name, _, _ in checks],
            "状态": ["已补齐" if ok else "待补" for _, ok, _ in checks],
            "说明": [desc for _, _, desc in checks],
        }
    )


def _gcard_summary_markdown_lines(*, train_dir: Path, eval_dir: Path, feature_dir: Path) -> list[str]:
    compare_score = "gcard_v6"
    compare_label = VERSION_LABELS.get(compare_score, compare_score)
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    overall = _read_csv(eval_dir / "overall_metrics.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    importance = _read_csv(train_dir / "feature_importance.csv")
    psi = _read_csv(eval_dir / "score_psi_by_month.csv")

    lines = [
        f"## Summary（新版模型 vs {compare_label}）",
        "",
        f"- 口径：新版模型（model_score） vs 旧版全客群模型（{compare_label}）；分客群表仅为切片效果，不含分客群专属模型对比。",
        f"- 更新日期：{time.strftime('%Y-%m-%d')}；指标来源：当前 run 已注册 train/evaluation artifacts。",
        "- 详情导航：模型描述、模型效果-每月效果、模型效果-模型sloping、模型稳定性、重要变量。",
        "",
        "### 一、结论摘要",
        "",
    ]
    lines.extend(_markdown_table(_gcard_summary_conclusion_frame(overall=overall, compare_score=compare_score)))
    lines.extend(
        [
            "",
            "### 二、模型与样本口径",
            "",
        ]
    )
    lines.extend(_markdown_table(_gcard_model_description_frame(train_dir=train_dir, feature_dir=feature_dir, run_config=run_config, metrics=metrics)))
    lines.extend(["", "### 三、整体效果对比", ""])
    lines.append("整体效果对比 AUC")
    lines.extend(_markdown_table(_gcard_metric_comparison_frame(overall, row_col="final_flag", metric="auc", row_label="样本", compare_score=compare_score)))
    lines.append("")
    lines.append("整体效果对比 KS")
    lines.extend(_markdown_table(_gcard_metric_comparison_frame(overall, row_col="final_flag", metric="ks", row_label="样本", compare_score=compare_score)))

    if segment is not None and not segment.empty:
        oot_oos_segment = _ordered_subset(segment, "segment", ["全客群", "老户次新", "老户", "次新", "流失户"])
        oot_oos_segment = oot_oos_segment[oot_oos_segment["final_flag"] == "OOT-OOS"] if "final_flag" in oot_oos_segment.columns else oot_oos_segment
        lines.extend(["", "### 四、OOT-OOS 分客群切片效果", "", "> 分客群结果是效果切片，不代表已训练老户次新/流失户专属模型。", ""])
        lines.append("OOT-OOS 分客群切片 AUC")
        lines.extend(_markdown_table(_gcard_metric_comparison_frame(oot_oos_segment, row_col="segment", metric="auc", row_label="客群", compare_score=compare_score, include_bad_rate=True)))
        lines.append("")
        lines.append("OOT-OOS 分客群切片 KS")
        lines.extend(_markdown_table(_gcard_metric_comparison_frame(oot_oos_segment, row_col="segment", metric="ks", row_label="客群", compare_score=compare_score, include_bad_rate=True)))

    if monthly is not None and not monthly.empty:
        oos_monthly = monthly[monthly["final_flag"].isin(["DEV-OOS", "OOT-OOS"])].copy() if "final_flag" in monthly.columns else monthly.copy()
        oos_monthly["_period_label"] = oos_monthly["final_flag"].astype(str) + " " + oos_monthly["mdl_month"].astype(str)
        lines.extend(["", "### 五、OOS 按月效果", ""])
        lines.append("OOS 按月 AUC")
        lines.extend(_markdown_table(_gcard_metric_comparison_frame(oos_monthly, row_col="_period_label", metric="auc", row_label="样本月份", compare_score=compare_score, include_bad_rate=True), limit=20))
        lines.append("")
        lines.append("OOS 按月 KS")
        lines.extend(_markdown_table(_gcard_metric_comparison_frame(oos_monthly, row_col="_period_label", metric="ks", row_label="样本月份", compare_score=compare_score, include_bad_rate=True), limit=20))

    lines.extend(["", "### 六、Sloping 高分10%表现", ""])
    lines.extend(_markdown_table(_gcard_top_decile_summary_frame(eval_dir=eval_dir, compare_score=compare_score)))

    if psi is not None and not psi.empty:
        lines.extend(["", "### 七、模型稳定性", "", "> PSI 为本轮模型分数月度汇总；分箱明细和 PSI component 见【模型稳定性】或待补口径说明。", ""])
        lines.extend(_markdown_table(_gcard_psi_summary_frame(psi=psi, compare_score=compare_score), limit=20))

    if importance is not None and not importance.empty:
        lines.extend(["", "### 八、Top 10 重要变量", "", "> 变量明细与 WOE 图见【重要变量】和【Top变量WOE】。", ""])
        lines.extend(_markdown_table(_gcard_top_features_frame(importance).head(10), limit=10))

    lines.extend(
        [
            "",
            "> 证据提示：本页不补造缺失结果；当前 run audit 仍应以 `rmw run audit` 输出为准。",
        ]
    )
    return lines


def _gcard_summary_conclusion_frame(*, overall: pd.DataFrame | None, compare_score: str) -> pd.DataFrame:
    row = _row_by_value(overall, "final_flag", "OOT-OOS")
    compare_label = VERSION_LABELS.get(compare_score, compare_score)
    if row:
        ks_diff = _to_float(row.get("model_score_ks")) - _to_float(row.get(f"{compare_score}_ks")) if None not in [_to_float(row.get("model_score_ks")), _to_float(row.get(f"{compare_score}_ks"))] else None
        auc_diff = _to_float(row.get("model_score_auc")) - _to_float(row.get(f"{compare_score}_auc")) if None not in [_to_float(row.get("model_score_auc")), _to_float(row.get(f"{compare_score}_auc"))] else None
        core = (
            f"OOT-OOS 全客群：本轮 KS {_fmt_metric(row.get('model_score_ks'))} vs {compare_label} {_fmt_metric(row.get(f'{compare_score}_ks'))}"
            f"，提升 {_fmt_pp(ks_diff)} 个百分点；AUC {_fmt_metric(row.get('model_score_auc'))} vs {compare_label} {_fmt_metric(row.get(f'{compare_score}_auc'))}"
            f"，提升 {_fmt_pp(auc_diff)} 个百分点。"
        )
    else:
        core = "OOT-OOS 全客群暂无可用对比指标。"
    return pd.DataFrame(
        [
            ("核心结论", core),
            ("对比对象", f"{compare_label} 作为旧版全客群模型；其余历史版本仍保留在明细 sheet 中供追溯。"),
            ("解释边界", "本页只聚焦 30天发起标签的 AUC/KS、OOS 月度表现、分客群切片、sloping、PSI；MOB/金额风险不是本页验收主口径。"),
        ],
        columns=["项目", "摘要"],
    )


def _gcard_top_decile_summary_frame(*, eval_dir: Path, compare_score: str) -> pd.DataFrame:
    rows = []
    for segment_label, segment_key in SEGMENT_FILES.items():
        model = _gcard_top_decile_stat(eval_dir=eval_dir, segment_key=segment_key, score_column="model_score")
        compare = _gcard_top_decile_stat(eval_dir=eval_dir, segment_key=segment_key, score_column=compare_score)
        if model is None and compare is None:
            continue
        rows.append(
            {
                "客群": segment_label,
                "本轮高分10%发起率": model.get("bad_rate") if model else None,
                f"{VERSION_LABELS.get(compare_score, compare_score)}高分10%发起率": compare.get("bad_rate") if compare else None,
                "发起率提升": (model.get("bad_rate") - compare.get("bad_rate")) if model and compare else None,
                "本轮累计lift": model.get("cum_lift") if model else None,
                f"{VERSION_LABELS.get(compare_score, compare_score)}累计lift": compare.get("cum_lift") if compare else None,
            }
        )
    return pd.DataFrame(rows)


def _gcard_top_decile_stat(*, eval_dir: Path, segment_key: str, score_column: str) -> dict[str, Any] | None:
    frame = _read_csv(eval_dir / f"decile_lift_{segment_key}_{score_column}.csv")
    if frame is None and score_column == "model_score":
        frame = _read_csv(eval_dir / f"decile_lift_{segment_key}.csv")
    if frame is None or frame.empty or "decile" not in frame.columns:
        return None
    top = frame[pd.to_numeric(frame["decile"], errors="coerce") == 10]
    if top.empty:
        return None
    row = top.iloc[0]
    return {"bad_rate": row.get("bad_rate"), "cum_lift": row.get("cum_lift")}


def _build_description_sheet(
    ws,
    *,
    train_dir: Path,
    eval_dir: Path,
    input_dir: Path,
    sample_dir: Path,
    feature_dir: Path,
) -> None:
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    eval_summary = _read_json(eval_dir / "evaluation_summary.json")
    feature_count = _feature_count(train_dir, feature_dir)

    row = 1
    row = _write_kv_section(
        ws,
            row,
            "报告信息",
            [
                ("报告名称", REPORT_TITLE),
                ("生成日期", time.strftime("%Y-%m-%d")),
                ("模型口径", "本轮基于新数据、新特征集合和新时间窗口训练主模型"),
                ("评估说明", "分客群结果作为效果切片评估，是否分客群训练由建模需求文档决定"),
        ],
    )

    row = _write_kv_section(
        ws,
        row,
        "样本与训练配置",
        [
            ("标签字段", run_config.get("label_column", "ftr_30d_ord_flag")),
            ("切分字段", run_config.get("split_column", "final_flag")),
            ("训练样本", f"final_flag in {run_config.get('train_values', [])}"),
            ("验证样本", f"final_flag in {run_config.get('valid_values', [])}"),
            ("OOS 样本", f"final_flag in {run_config.get('oos_values', [])}"),
            ("算法", run_config.get("algorithm", "N/A")),
            ("入模特征数", feature_count or run_config.get("actual_feature_count", "N/A")),
            ("Best iteration", run_config.get("best_iteration", "N/A")),
            ("随机种子", run_config.get("random_seed", "N/A")),
        ],
    )

    row = _write_kv_section(
        ws,
        row,
        "Train / Valid 核心效果",
        [
            ("Train 样本数", metrics.get("train_samples", run_config.get("train_samples", "N/A"))),
            ("Valid 样本数", metrics.get("valid_samples", run_config.get("valid_samples", "N/A"))),
            ("Train bad rate", metrics.get("train_bad_rate", run_config.get("train_bad_rate", "N/A"))),
            ("Valid bad rate", metrics.get("valid_bad_rate", run_config.get("valid_bad_rate", "N/A"))),
            ("Train AUC", metrics.get("train_auc", "N/A")),
            ("Valid AUC", metrics.get("valid_auc", "N/A")),
            ("Train KS", metrics.get("train_ks", "N/A")),
            ("Valid KS", metrics.get("valid_ks", "N/A")),
            ("AUC Gap", metrics.get("auc_gap", "N/A")),
            ("训练耗时秒", metrics.get("train_time_seconds", "N/A")),
        ],
    )

    if eval_summary:
        row = _write_kv_section(
            ws,
            row,
            "评估范围",
            [
                ("总样本数", eval_summary.get("n_total_samples", "N/A")),
                ("评估分数", ", ".join(eval_summary.get("score_columns_evaluated", []))),
                ("评估切分", ", ".join(eval_summary.get("splits_evaluated", []))),
            ],
        )

    for title, path in [
        ("样本切分分布", _first_existing(sample_dir / "sample_split_summary.csv", input_dir / "sample_split_summary.csv")),
        ("标签分布", _first_existing(sample_dir / "label_distribution.csv", input_dir / "label_distribution.csv")),
        ("客群分布", _first_existing(sample_dir / "segment_distribution.csv", input_dir / "segment_distribution.csv")),
    ]:
        frame = _read_csv(path) if path else None
        if frame is not None and not frame.empty:
            row = _write_table(ws, row, title, frame)

    row = _write_kv_section(
        ws,
        row,
        "口径提示",
        [
            ("风险观察", "MOB1/MOB3 历史风险定义仍需确认，当前风险表不等同于历史正式口径"),
            ("缺失补充", "详见 reports/model_report_missing_results.md"),
        ],
    )


def _build_features_sheet(ws, *, train_dir: Path, feature_dir: Path) -> None:
    importance = _read_csv(train_dir / "feature_importance.csv")
    drop_detail = _read_csv(train_dir / "feature_drop_detail.csv")
    availability = _read_csv(feature_dir / "feature_availability.csv")
    final_features = _read_feature_list(train_dir, feature_dir)

    row = 1
    if importance is not None and not importance.empty:
        enriched = importance.copy()
        if drop_detail is not None and not drop_detail.empty:
            keep_cols = [col for col in ["feature", "non_null_rate", "unique_count", "drop_reason"] if col in drop_detail.columns]
            enriched = enriched.merge(drop_detail[keep_cols], on="feature", how="left")
        if availability is not None and not availability.empty:
            keep_cols = [col for col in ["feature", "in_feather", "dtype"] if col in availability.columns]
            enriched = enriched.merge(availability[keep_cols], on="feature", how="left")
        row = _write_table(ws, row, "Top 20 重要变量", enriched.head(20))
        row = _write_table(ws, row, "完整入模变量明细", enriched)

    if final_features:
        feature_frame = pd.DataFrame({"index": range(1, len(final_features) + 1), "feature": final_features})
        row = _write_table(ws, row, "最终特征清单", feature_frame)

    if drop_detail is not None and "drop_reason" in drop_detail.columns:
        dropped = drop_detail[drop_detail["drop_reason"].notna() & (drop_detail["drop_reason"].astype(str) != "")]
        if not dropped.empty:
            _write_table(ws, row, "被删除特征", dropped)


def _build_woe_sheet(ws, *, train_dir: Path, report_dir: Path) -> None:
    summary_path = _find_woe_summary(train_dir=train_dir, report_dir=report_dir)
    if summary_path is None:
        _write_note(ws, 1, "WOE charts require row-level feature values. No registered Top feature WOE artifacts were found.")
        return

    summary = _read_csv(summary_path)
    if summary is None or summary.empty:
        _write_note(ws, 1, "WOE charts require row-level feature values. The WOE summary artifact is empty.")
        return

    ok = summary[summary.get("status", "") == "ok"].copy()
    if ok.empty:
        row = _write_note(ws, 1, "WOE charts require row-level feature values. No Top feature produced a renderable WOE chart.")
        _write_table(ws, row, "WOE skipped features", summary[[col for col in ["feature", "rank", "status", "skip_reason"] if col in summary.columns]], apply_color_scale=False)
        return

    row = 1
    image_dir = summary_path.parent / "images"
    for (_, feature), feature_rows in ok.groupby(["rank", "feature"], sort=True):
        rank = int(feature_rows["rank"].iloc[0])
        gain = feature_rows["gain"].iloc[0]
        base_split = "DEV" if "DEV" in set(feature_rows["split_value"].astype(str)) else str(feature_rows["split_value"].iloc[0])
        base_rows = feature_rows[feature_rows["split_value"].astype(str) == base_split]
        iv = base_rows.groupby("bin_label", as_index=False)["iv_component"].first()["iv_component"].sum()
        missing_rate = base_rows.loc[base_rows["is_missing_bin"].astype(bool), "pop_pct"].sum() if "is_missing_bin" in base_rows else 0
        image_path = _find_woe_image(image_dir, rank)
        row = _write_kv_section(
            ws,
            row,
            f"Top {rank}: {feature}",
            [
                ("Gain", gain),
                ("Base split", base_split),
                ("IV", iv),
                ("Missing rate", missing_rate),
                ("Image", image_path.name if image_path else "missing"),
            ],
        )
        if image_path:
            try:
                from openpyxl.drawing.image import Image as OpenpyxlImage

                image = OpenpyxlImage(str(image_path))
                image.width = 900
                image.height = 420
                ws.add_image(image, f"A{row}")
                row += 24
            except Exception as exc:
                row = _write_note(ws, row, f"WOE image could not be embedded: {image_path.name}; {exc}")
        else:
            row = _write_note(ws, row, f"WOE image missing for Top {rank}: {feature}")


def _build_screening_params_sheet(ws, *, train_dir: Path, feature_dir: Path) -> None:
    stage_summary = _read_json(feature_dir / "feature_stage_summary.json")
    run_config = _read_json(train_dir / "run_config.json")
    params = run_config.get("params", {}) if isinstance(run_config.get("params"), dict) else {}

    row = 1
    if stage_summary:
        row = _write_table(ws, row, "变量筛选过程", _screening_steps_frame(stage_summary, feature_dir))
        row = _write_kv_section(
            ws,
            row,
            "筛选结果摘要",
            [
                ("筛选方法", stage_summary.get("filtering_method", "N/A")),
                ("数据可用特征数", stage_summary.get("features_available_in_data", "N/A")),
                ("缺失特征数", stage_summary.get("missing_features", "N/A")),
                ("潜在泄露提示", stage_summary.get("potential_leakage_flags", "无")),
            ],
        )
    else:
        row = _write_table(ws, row, "训练特征准备", _training_feature_preparation_frame(train_dir, run_config))

    if params:
        row = _write_table(ws, row, "LightGBM 参数", pd.DataFrame({"parameter": list(params.keys()), "value": list(params.values())}))

    _write_kv_section(
        ws,
        row,
        "训练配置",
        [
            ("experiment", run_config.get("experiment", "N/A")),
            ("data_source", run_config.get("data_source", "N/A")),
            ("candidate_feature_count", run_config.get("candidate_feature_count", "N/A")),
            ("actual_feature_count", run_config.get("actual_feature_count", "N/A")),
        ],
    )


def _model_conclusion_summary(eval_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    monthly_oos = _read_csv(eval_dir / "monthly_segment_metrics_oos_by_version.csv")
    decile = _read_csv(eval_dir / "decile_lift_bins_by_version.csv")
    ftr_rate = _read_csv(eval_dir / "intent_zc_segment_ftr_rate_by_version.csv")
    amount_risk = _read_csv(eval_dir / "intent_zc_segment_amount_risk_by_version.csv")

    def add(module: str, conclusion: str) -> None:
        if conclusion:
            rows.append({"模块": module, "结论": conclusion})

    comparison_scores = _comparison_score_columns()
    primary_comparison_scores = comparison_scores[-2:] if len(comparison_scores) >= 2 else comparison_scores

    if segment is not None and not segment.empty:
        for idx, segment_name in enumerate(["老户次新", "流失户"], start=1):
            matched = segment[(segment["segment"] == segment_name) & (segment["final_flag"] == "OOT-OOS")]
            if matched.empty:
                continue
            row = matched.iloc[0].to_dict()
            compare_text = "，".join(_ks_compare_text(row, version) for version in primary_comparison_scores)
            if not compare_text:
                compare_text = "未配置可比 champion 分数"
            add(
                "1、每月效果（OOS）",
                f"（{idx}）在 OOT-OOS 样本上{segment_name}客群上，本轮全客群模型 KS "
                f"{_fmt_metric(row.get('model_score_ks'))}；{compare_text}。",
            )
        overall = segment[(segment["segment"] == "全客群") & (segment["final_flag"] == "OOT-OOS")]
        if not overall.empty:
            row = overall.iloc[0].to_dict()
            overall_compare_text = "，".join(_ks_compare_text(row, version) for version in primary_comparison_scores)
            if not overall_compare_text:
                overall_compare_text = "未配置可比 champion 分数"
            add(
                "1、每月效果（OOS）",
                "（3）在 OOT-OOS 样本上全客群整体看，本轮模型 KS "
                f"{_fmt_metric(row.get('model_score_ks'))}，{overall_compare_text}。",
            )
        if primary_comparison_scores:
            labels = "/".join(VERSION_LABELS.get(score, score) for score in primary_comparison_scores)
            add(
                "1、每月效果（OOS）",
                "（4）当前 run 未注册老户次新/流失户专属模型得分，无法复刻历史文档中的“分客群建模 KS”对比；"
                f"本摘要仅比较本轮全客群模型与已注册的 {labels} 历史版本。",
            )

    if monthly_oos is not None and not monthly_oos.empty:
        month_texts = []
        for segment_name in ["老户次新", "流失户"]:
            subset = monthly_oos[
                (monthly_oos["segment"] == segment_name)
                & (monthly_oos["final_flag"] == "OOT-OOS")
                & (monthly_oos["score_version"] == "model_score")
            ].sort_values("mdl_month")
            if subset.empty:
                continue
            pieces = [f"{row.mdl_month} KS {_fmt_metric(row.ks)}" for row in subset.itertuples(index=False)]
            month_texts.append(f"{segment_name}：" + "、".join(pieces))
        if month_texts:
            add(
                "1、每月效果（OOS）",
                "（5）DEV-OOS 与 OOT-OOS 已拼接到【模型效果-每月效果】中横向比较；"
                f"OOT-OOS 本轮模型 by 月结果为：{'；'.join(month_texts)}。",
            )

    if decile is not None and not decile.empty:
        for idx, segment_name in enumerate(["老户次新", "流失户"], start=1):
            model_stat = _top_decile_stat(decile, segment_name, "model_score")
            comparison_stats = [(score, _top_decile_stat(decile, segment_name, score)) for score in primary_comparison_scores]
            comparison_stats = [(score, stat) for score, stat in comparison_stats if stat is not None]
            if model_stat is None or not comparison_stats:
                continue
            compare_text = "，".join(
                f"对比{VERSION_LABELS.get(score, score)}为{_fmt_percent(stat['bad_rate'])}、lift {_fmt_metric(stat['lift'])}"
                for score, stat in comparison_stats
            )
            add(
                "2、模型sloping",
                f"（{idx}）在 OOT-OOS 样本上{segment_name}客群高分10%分层，本轮模型30天发起率"
                f"{_fmt_percent(model_stat['bad_rate'])}、lift {_fmt_metric(model_stat['lift'])}；"
                f"{compare_text}。",
            )

    if ftr_rate is not None and amount_risk is not None and not ftr_rate.empty and not amount_risk.empty:
        add("3、意愿交叉风险（DEV-OOS）", "（1）高、中、低意愿评级为对应模型分数在各客群内三等频分箱得到。")
        for idx, segment_name in enumerate(["老户", "流失户"], start=2):
            model_low_ftr = _intent_total_value(ftr_rate, segment_name, "低意愿", "ftr_30d_rate", "model_score")
            model_high_ftr = _intent_total_value(ftr_rate, segment_name, "高意愿", "ftr_30d_rate", "model_score")
            model_high_risk = _intent_total_value(amount_risk, segment_name, "高意愿", "amount_overdue_rate", "model_score")
            comparison_values = []
            for score in primary_comparison_scores:
                low_ftr = _intent_total_value(ftr_rate, segment_name, "低意愿", "ftr_30d_rate", score)
                high_ftr = _intent_total_value(ftr_rate, segment_name, "高意愿", "ftr_30d_rate", score)
                high_risk = _intent_total_value(amount_risk, segment_name, "高意愿", "amount_overdue_rate", score)
                if None not in [low_ftr, high_ftr, high_risk]:
                    comparison_values.append((score, low_ftr, high_ftr, high_risk))
            if None in [model_low_ftr, model_high_ftr, model_high_risk] or not comparison_values:
                continue
            ftr_compare = "，".join(
                f"对比{VERSION_LABELS.get(score, score)}低/高意愿为{_fmt_percent(low_ftr)}/{_fmt_percent(high_ftr)}"
                for score, low_ftr, high_ftr, _ in comparison_values
            )
            risk_compare = "，".join(
                f"{VERSION_LABELS.get(score, score)}为{_fmt_percent(high_risk)}"
                for score, _, _, high_risk in comparison_values
            )
            add(
                "3、意愿交叉风险（DEV-OOS）",
                f"（{idx}）{segment_name}客群上，本轮低意愿30天发起率{_fmt_percent(model_low_ftr)}、"
                f"高意愿30天发起率{_fmt_percent(model_high_ftr)}；"
                f"{ftr_compare}；高意愿新增订单3期金额逾期率本轮{_fmt_percent(model_high_risk)}，"
                f"{risk_compare}。",
            )

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame({"序号": range(1, len(rows) + 1), "模块": [row["模块"] for row in rows], "结论": [row["结论"] for row in rows]})


def _module_conclusion_frame(eval_dir: Path, module: str) -> pd.DataFrame:
    summary = _model_conclusion_summary(eval_dir)
    if summary.empty:
        return pd.DataFrame()
    subset = summary[summary["模块"] == module].copy()
    if subset.empty:
        return pd.DataFrame()
    subset["序号"] = range(1, len(subset) + 1)
    return subset[["序号", "结论"]]


def _write_module_conclusions(ws, row: int, eval_dir: Path, module: str) -> int:
    frame = _module_conclusion_frame(eval_dir, module)
    if frame.empty:
        return row
    return _write_table(ws, row, module, frame, apply_color_scale=False)


def _ks_compare_text(row: dict[str, Any], score_column: str) -> str:
    label = VERSION_LABELS.get(score_column, score_column)
    model_ks = _to_float(row.get("model_score_ks"))
    version_ks = _to_float(row.get(f"{score_column}_ks"))
    if model_ks is None or version_ks is None:
        return f"{label} KS 暂无可比数据"
    diff = model_ks - version_ks
    if diff >= 0:
        return f"对比{label} KS {_fmt_metric(version_ks)} 提升{_fmt_pp(diff)}个百分点"
    return f"较{label} KS {_fmt_metric(version_ks)} 低{_fmt_pp(abs(diff))}个百分点"


def _comparison_score_columns() -> list[str]:
    return [score for score in SCORE_COLUMNS if score != "model_score"]


def _top_decile_stat(frame: pd.DataFrame, segment_name: str, score_column: str) -> dict[str, float] | None:
    subset = frame[
        (frame["segment"] == segment_name)
        & (frame["final_flag"] == "OOT-OOS")
        & (frame["score_version"] == score_column)
    ].copy()
    if subset.empty:
        return None
    total_n = pd.to_numeric(subset["n_samples"], errors="coerce").sum()
    total_bad = pd.to_numeric(subset["bad"], errors="coerce").sum()
    top = subset[subset["decile"] == 10]
    if not total_n or not total_bad or top.empty:
        return None
    top_rate = _to_float(top.iloc[0]["bad_rate"])
    if top_rate is None:
        return None
    base_rate = total_bad / total_n
    return {"bad_rate": top_rate, "lift": top_rate / base_rate if base_rate else 0.0}


def _intent_total_value(
    frame: pd.DataFrame,
    segment_name: str,
    intent_level: str,
    value_col: str,
    score_column: str = "model_score",
) -> float | None:
    subset = frame[
        (frame["segment"] == segment_name)
        & (frame["final_flag"] == "DEV-OOS")
        & (frame["score_version"] == score_column)
        & (frame["intent_level"] == intent_level)
        & (frame["zc_level"].astype(str) == "合计")
    ]
    if subset.empty or value_col not in subset.columns:
        return None
    return _to_float(subset.iloc[0][value_col])


def _to_float(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _build_monthly_effect_sheet(ws, *, eval_dir: Path) -> None:
    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    benchmark = _read_csv(eval_dir / "benchmark_uplift.csv")
    monthly_segment_oos_versioned = _read_csv(eval_dir / "monthly_segment_metrics_oos_by_version.csv")
    monthly_segment_versioned = _read_csv(eval_dir / "monthly_segment_metrics_oot_oos_by_version.csv")
    row = 1
    row = _write_module_conclusions(ws, row, eval_dir, "1、每月效果（OOS）")

    if monthly is not None and not monthly.empty:
        for final_flag in _ordered_values(monthly["final_flag"].dropna().unique().tolist()):
            subset = monthly[monthly["final_flag"] == final_flag].copy()
            if subset.empty:
                continue
            title = f"在全客群 {final_flag} 样本效果"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame(subset, row_col="mdl_month", metric="ks", row_label="30天发起Y"),
                right_frame=_metric_comparison_frame(subset, row_col="mdl_month", metric="auc", row_label="30天发起Y"),
            )

    if monthly_segment_oos_versioned is not None and not monthly_segment_oos_versioned.empty:
        row = _write_note(
            ws,
            row,
            "以下为 OOS 客群 by 月效果，DEV-OOS 与 OOT-OOS 按时间顺序拼接，按同一客群横向对比本轮模型与历史版本。",
        )
        for segment_name in ["全客群", "老户次新", "老户", "次新", "流失户"]:
            subset = monthly_segment_oos_versioned[
                (monthly_segment_oos_versioned["segment"] == segment_name)
                & (monthly_segment_oos_versioned["final_flag"].isin(["DEV-OOS", "OOT-OOS"]))
            ].copy()
            if subset.empty:
                continue
            title = f"在{segment_name} OOS by月效果（DEV-OOS + OOT-OOS）"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame_oos_by_month(subset, metric="ks", row_label="样本月份"),
                right_frame=_metric_comparison_frame_oos_by_month(subset, metric="auc", row_label="样本月份"),
            )
    elif monthly_segment_versioned is not None and not monthly_segment_versioned.empty:
        row = _write_note(ws, row, "以下为 OOT-OOS 客群 by 月效果，按同一客群横向对比本轮模型与历史版本。")
        for segment_name in ["老户次新", "流失户"]:
            subset = monthly_segment_versioned[
                (monthly_segment_versioned["segment"] == segment_name)
                & (monthly_segment_versioned["final_flag"] == "OOT-OOS")
            ].copy()
            if subset.empty:
                continue
            title = f"在{segment_name} OOT-OOS by月效果"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame_long(subset, row_col="mdl_month", metric="ks", row_label="30天发起Y"),
                right_frame=_metric_comparison_frame_long(subset, row_col="mdl_month", metric="auc", row_label="30天发起Y"),
            )

    if segment is not None and not segment.empty:
        row = _write_note(ws, row, "以下分客群表为切片效果，不代表已训练分客群专属模型；每张小表只横向对比一个指标。")
        for segment_name in ["老户次新", "老户", "次新", "流失户", "全客群"]:
            subset = segment[segment["segment"] == segment_name].copy()
            if subset.empty:
                continue
            title = f"在{segment_name}效果"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame(subset, row_col="final_flag", metric="ks", row_label="样本"),
                right_frame=_metric_comparison_frame(subset, row_col="final_flag", metric="auc", row_label="样本"),
            )

    if benchmark is not None and not benchmark.empty:
        row = _write_table(ws, row, "历史版本提升摘要", benchmark)


def _build_sloping_sheet(ws, *, eval_dir: Path) -> None:
    versioned = _read_csv(eval_dir / "decile_lift_bins_by_version.csv")
    if versioned is not None and not versioned.empty:
        row = 1
        row = _write_module_conclusions(ws, row, eval_dir, "2、模型sloping")
        _build_sloping_sheet_from_versioned(ws, versioned, start_row=row)
        return

    row = 1
    row = _write_module_conclusions(ws, row, eval_dir, "2、模型sloping")
    for segment_label, segment_key in SEGMENT_FILES.items():
        row = _write_note(ws, row, f"2025-12 至 2026-01 30天发起 OOT/OOS：在{segment_label}效果")
        table_row = row
        max_end = row
        col = 1
        for score_column in SCORE_COLUMNS:
            frame = _read_csv(eval_dir / f"decile_lift_{segment_key}_{score_column}.csv")
            if frame is None and score_column == "model_score":
                frame = _read_csv(eval_dir / f"decile_lift_{segment_key}.csv")
            if frame is None or frame.empty:
                continue
            end = _write_table(
                ws,
                table_row,
                VERSION_LABELS.get(score_column, score_column),
                _sloping_display_frame(frame),
                start_col=col,
                apply_color_scale=False,
                plain=True,
            )
            max_end = max(max_end, end)
            col = 8 if col == 1 else 1
            if col == 1:
                table_row = max_end
        row = max(max_end + 1, table_row + 1)


def _build_sloping_sheet_from_versioned(ws, versioned: pd.DataFrame, *, start_row: int = 1) -> None:
    row = start_row
    for segment_label in ["全客群", "老户次新", "流失户"]:
        subset_segment = versioned[(versioned["segment"] == segment_label) & (versioned["final_flag"] == "OOT-OOS")].copy()
        if subset_segment.empty:
            continue
        row = _write_note(ws, row, f"OOT-OOS 30天发起：在{segment_label}效果")
        table_row = row
        max_end = row
        col = 1
        for score_column in SCORE_COLUMNS:
            subset = subset_segment[subset_segment["score_version"] == score_column].copy()
            if subset.empty:
                continue
            end = _write_table(
                ws,
                table_row,
                VERSION_LABELS.get(score_column, score_column),
                _sloping_display_frame(subset),
                start_col=col,
                apply_color_scale=False,
                plain=True,
            )
            max_end = max(max_end, end)
            col = 8 if col == 1 else 1
            if col == 1:
                table_row = max_end
        row = max(max_end + 1, table_row + 1)


def _build_intent_risk_sheet(ws, *, eval_dir: Path) -> None:
    distribution_by_version = _read_csv(eval_dir / "intent_zc_segment_distribution_by_version.csv")
    ftr_by_version = _read_csv(eval_dir / "intent_zc_segment_ftr_rate_by_version.csv")
    amount_by_version = _read_csv(eval_dir / "intent_zc_segment_amount_risk_by_version.csv")
    if (
        distribution_by_version is not None
        and not distribution_by_version.empty
        and ftr_by_version is not None
        and not ftr_by_version.empty
        and amount_by_version is not None
        and not amount_by_version.empty
    ):
        row = 1
        row = _write_module_conclusions(ws, row, eval_dir, "3、意愿交叉风险（DEV-OOS）")
        _build_intent_risk_sheet_from_versioned(ws, distribution_by_version, ftr_by_version, amount_by_version, start_row=row)
        return

    row = 1
    row = _write_module_conclusions(ws, row, eval_dir, "3、意愿交叉风险（DEV-OOS）")
    row = _write_note(
        ws,
        row,
        "当前意愿资产交叉 artifact 缺少老户/流失户、score version、final_flag 和金额风险 x 资产评级维度；下方仅展示当前可用的全量观察口径，待补口径见 missing 文档。",
    )
    distribution = _read_csv(eval_dir / "intent_zc_distribution.csv")
    amount_risk = _read_csv(eval_dir / "intent_zc_amount_risk.csv")
    head_risk = _read_csv(eval_dir / "intent_zc_headcount_risk.csv")

    if distribution is not None and not distribution.empty:
        row = _write_table(ws, row, "当前可用全量观察：占比（意愿评级 x 资产评级）", _intent_sum_matrix(distribution, "pct"))
        row = _write_table(ws, row, "当前可用全量观察：30天发起率（意愿评级 x 资产评级）", _intent_rate_matrix(distribution, "bad", "n_samples"))
    if head_risk is not None and not head_risk.empty:
        row = _write_table(ws, row, "当前可用全量观察：人头风险率（意愿评级 x 资产评级）", _intent_rate_matrix(head_risk, "head_risk_count", "n_samples"))
    if amount_risk is not None and not amount_risk.empty:
        display = amount_risk.rename(
            columns={
                "intent_level": "意愿",
                "n_samples": "样本数",
                "total_principal": "本金金额",
                "total_overdue": "逾期金额",
                "amount_overdue_rate": "金额逾期率",
                "head_risk_count": "人头风险数",
                "head_risk_rate": "人头风险率",
            }
        )
        row = _write_table(ws, row, "当前可用全量观察：金额风险（仅意愿维度）", display)

    target = pd.DataFrame(
        [
            ("老户", "占比", "资产评级 x 意愿评级，含行/列 sum"),
            ("老户", "30天发起率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("老户", "新增订单3期金额逾期率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("流失户", "占比", "资产评级 x 意愿评级，含行/列 sum"),
            ("流失户", "30天发起率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("流失户", "新增订单3期金额逾期率", "资产评级 x 意愿评级，含行/列加权整体"),
        ],
        columns=["待补客群", "待补指标", "目标矩阵"],
    )
    _write_table(ws, row, "待补矩阵口径", target, apply_color_scale=False)


def _build_intent_risk_sheet_from_versioned(
    ws,
    distribution: pd.DataFrame,
    ftr_rate: pd.DataFrame,
    amount_risk: pd.DataFrame,
    *,
    start_row: int = 1,
) -> None:
    row = start_row
    row = _write_note(
        ws,
        row,
        "以下矩阵均限定 DEV-OOS；意愿评级按对应客群、对应分数版本等频三等分，资产评级为 zc_level 1-7，并保留行列合计。",
    )
    matrix_specs = [
        ("占比", distribution, "sample_pct", True),
        ("30天发起率", ftr_rate, "ftr_30d_rate", True),
        ("新增订单3期金额逾期率", amount_risk, "amount_overdue_rate", True),
    ]
    for segment_name in ["老户", "流失户"]:
        row = _write_note(ws, row, f"{segment_name} DEV-OOS 意愿 x 资产评级")
        for metric_label, frame, value_col, heatmap in matrix_specs:
            metric_frame = frame[(frame["segment"] == segment_name) & (frame["final_flag"] == "DEV-OOS")].copy()
            if metric_frame.empty:
                continue
            for score_column in SCORE_COLUMNS:
                subset = metric_frame[metric_frame["score_version"] == score_column].copy()
                if subset.empty:
                    continue
                title = f"{segment_name} - {metric_label} - {VERSION_LABELS.get(score_column, score_column)}"
                row = _write_table(
                    ws,
                    row,
                    title,
                    _intent_version_matrix(subset, value_col),
                    color_scale_data=heatmap,
                    color_scale_prefer_high=True,
                )


def _build_stability_sheet(ws, *, eval_dir: Path) -> None:
    distribution = _read_csv(eval_dir / "model_score_bin_distribution_by_month.csv")
    if distribution is not None and not distribution.empty:
        _build_stability_sheet_from_distribution(ws, distribution)
        return

    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    row = 1
    if psi is None or psi.empty:
        return

    if "score_column" in psi.columns:
        model_psi = psi[psi["score_column"] == "model_score"].copy()
        if model_psi.empty:
            model_psi = psi.copy()
    else:
        model_psi = psi.copy()

    display = model_psi.rename(columns={"month": "月份", "psi": "PSI", "n_samples": "样本数", "score_column": "分数"})
    row = _write_table(ws, row, "本轮模型月度 PSI", display)
    row = _write_note(ws, row, "当前 artifact 只有本轮模型月度 PSI 汇总，缺少每个分箱的跨月占比和发起率变化；待补口径见 missing 文档。")
    target = pd.DataFrame(
        [
            ("model_score", "月份 x 分箱", "score_bin/decile", "样本量、占比、30天发起率、PSI component"),
            ("model_score", "基准月 vs 观察月", "score_bin/decile", "占比差异、发起率差异"),
        ],
        columns=["分数", "期望粒度", "分箱字段", "期望指标"],
    )
    _write_table(ws, row, "待补稳定性分箱明细口径", target, apply_color_scale=False)


def _build_stability_sheet_from_distribution(ws, distribution: pd.DataFrame) -> None:
    row = 1
    display = distribution.copy()
    if "score_column" in display.columns:
        display = display[display["score_column"] == "model_score"].copy()
    display = display.sort_values(["score_decile", "mdl_month"])

    pct_pivot = _stability_pivot(display, "pct", "占比")
    row = _write_table(
        ws,
        row,
        "本轮模型分箱占比变化",
        pct_pivot,
        color_scale_data=True,
        color_scale_prefer_high=True,
        color_scale_start_offset=2,
    )

    bad_rate_pivot = _stability_pivot(display, "bad_rate", "30天发起率")
    row = _write_table(
        ws,
        row,
        "本轮模型分箱30天发起率变化",
        bad_rate_pivot,
        color_scale_data=True,
        color_scale_prefer_high=True,
        color_scale_start_offset=2,
    )

    psi_cols = ["mdl_month", "month_psi"]
    if "n_samples" in display.columns:
        psi_summary = display.groupby("mdl_month", as_index=False).agg({"n_samples": "sum", "month_psi": "first"})
    else:
        psi_summary = display[psi_cols].drop_duplicates().copy()
    psi_summary = psi_summary.rename(columns={"mdl_month": "月份", "n_samples": "样本数", "month_psi": "PSI"})
    row = _write_table(ws, row, "本轮模型月度 PSI", psi_summary)

    detail_cols = [
        "mdl_month",
        "score_decile",
        "lower_bound",
        "n_samples",
        "pct",
        "bad_rate",
        "baseline_pct",
        "psi_component",
        "month_psi",
    ]
    detail = display[[col for col in detail_cols if col in display.columns]].rename(
        columns={
            "mdl_month": "月份",
            "score_decile": "分箱",
            "lower_bound": "分组",
            "n_samples": "样本数",
            "pct": "占比",
            "bad_rate": "30天发起率",
            "baseline_pct": "基准占比",
            "psi_component": "PSI组件",
            "month_psi": "月度PSI",
        }
    )
    _write_table(ws, row, "本轮模型稳定性分箱明细", detail)


def _write_table(
    ws,
    start_row: int,
    title: str,
    frame: pd.DataFrame,
    *,
    source_note: str | None = None,
    start_col: int = 1,
    apply_color_scale: bool = True,
    plain: bool = False,
    color_scale_data: bool = False,
    color_scale_prefer_high: bool = False,
    color_scale_start_offset: int = 1,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    frame = _clean_frame(frame)
    if frame.empty:
        return start_row

    ncols = max(len(frame.columns), 1)
    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=start_col + ncols - 1)
    title_cell = ws.cell(row=title_row, column=start_col)
    title_cell.value = title
    title_cell.font = Font(name="楷体", size=12, bold=True)
    if not plain:
        title_cell.fill = PatternFill(start_color="E7F4F2", end_color="E7F4F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = title_row + 1
    if source_note:
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col + ncols - 1)
        note_cell = ws.cell(row=header_row, column=start_col)
        note_cell.value = source_note
        note_cell.font = Font(name="楷体", size=10, italic=True, color="666666")
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_row += 1

    for col_offset, col_name in enumerate(frame.columns):
        cell = ws.cell(row=header_row, column=start_col + col_offset)
        cell.value = col_name
        cell.font = Font(name="楷体", size=10, bold=True)
        if not plain:
            cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row_data in enumerate(frame.itertuples(index=False), 1):
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=header_row + row_offset, column=start_col + col_offset)
            column_name = str(frame.columns[col_offset])
            cell.value = _excel_value(value, column_name)
            cell.number_format = _number_format(column_name, cell.value)
            cell.alignment = _alignment(column_name)

    end_row = header_row + len(frame)
    _style_region(ws, title_row, end_row, ncols, start_col=start_col)
    if apply_color_scale:
        _apply_table_color_scales(ws, header_row, end_row, frame, start_col=start_col)
    if color_scale_data and ncols > color_scale_start_offset:
        _apply_color_scale(
            ws,
            header_row + 1,
            end_row,
            start_col + color_scale_start_offset,
            start_col + ncols - 1,
            prefer_high=color_scale_prefer_high,
        )
    _update_column_widths(ws, frame, ncols, start_col=start_col)

    for col_idx in range(start_col, start_col + ncols):
        ws.column_dimensions[get_column_letter(col_idx)].bestFit = False

    return end_row + 3


def _write_metric_table_pair(ws, start_row: int, *, title: str, left_frame: pd.DataFrame, right_frame: pd.DataFrame) -> int:
    left_end = _write_comparison_table(ws, start_row, title, "KS", left_frame, start_col=1)
    right_end = _write_comparison_table(ws, start_row, title, "AUC", right_frame, start_col=9)
    return max(left_end, right_end)


def _write_comparison_table(
    ws,
    start_row: int,
    title: str,
    metric_label: str,
    frame: pd.DataFrame,
    *,
    start_col: int,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    frame = _clean_frame(frame)
    ncols = len(frame.columns)
    title_row = start_row
    metric_row = title_row + 1
    header_row = title_row + 2

    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=start_col + ncols - 1)
    title_cell = ws.cell(title_row, start_col)
    title_cell.value = title
    title_cell.font = Font(name="楷体", size=12, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=metric_row, start_column=start_col + 1, end_row=metric_row, end_column=start_col + ncols - 1)
    ws.cell(metric_row, start_col).value = frame.columns[0]
    ws.cell(metric_row, start_col).font = Font(name="楷体", size=10, bold=True)
    ws.cell(metric_row, start_col).alignment = Alignment(horizontal="center", vertical="center")
    metric_cell = ws.cell(metric_row, start_col + 1)
    metric_cell.value = metric_label
    metric_cell.font = Font(name="楷体", size=11, bold=True)
    metric_cell.alignment = Alignment(horizontal="center", vertical="center")
    metric_cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")

    for col_offset, col_name in enumerate(frame.columns):
        cell = ws.cell(header_row, start_col + col_offset)
        cell.value = col_name if col_offset > 0 else ""
        cell.font = Font(name="楷体", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")

    for row_offset, row_data in enumerate(frame.itertuples(index=False), 1):
        for col_offset, value in enumerate(row_data):
            column_name = str(frame.columns[col_offset])
            cell = ws.cell(header_row + row_offset, start_col + col_offset)
            cell.value = _excel_value(value, column_name)
            cell.number_format = _number_format(column_name, cell.value)
            cell.alignment = _alignment(column_name)

    end_row = header_row + len(frame)
    _style_region(ws, title_row, end_row, ncols, start_col=start_col)
    _apply_color_scale(ws, header_row + 1, end_row, start_col + 1, start_col + ncols - 1, prefer_high=True)
    _update_column_widths(ws, frame, ncols, start_col=start_col)
    return end_row + 3


def _write_kv_section(ws, start_row: int, title: str, pairs: list[tuple[str, Any]]) -> int:
    frame = pd.DataFrame(pairs, columns=["项目", "内容"])
    return _write_table(ws, start_row, title, frame)


def _write_note(ws, start_row: int, text: str) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    cell = ws.cell(row=start_row, column=1)
    cell.value = text
    cell.font = Font(name="楷体", size=10, italic=True, color="9C6500")
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _style_region(ws, start_row, start_row, 8)
    return start_row + 2


def _find_woe_summary(*, train_dir: Path, report_dir: Path | None = None) -> Path | None:
    candidates = []
    if report_dir is not None:
        candidates.append(report_dir / "woe_top_features")
    candidates.append(train_dir / "woe_top_features")
    for directory in candidates:
        for path in sorted(directory.glob("woe_top*_summary.csv")):
            if path.exists():
                return path
    return None


def _find_woe_image(image_dir: Path, rank: int) -> Path | None:
    if not image_dir.exists():
        return None
    matches = sorted(image_dir.glob(f"{rank:03d}_*_WOE.png"))
    return matches[0] if matches else None


def _screening_steps_frame(stage_summary: dict[str, Any], feature_dir: Path) -> pd.DataFrame:
    process = _read_feature_screening_process(feature_dir, allow_legacy_fallback=_legacy_screening_process_allowed(feature_dir))
    if process:
        rows = process.get("screening_rows", [])
        if rows:
            return pd.DataFrame(
                {
                    "步骤": [row.get("step") for row in rows],
                    "筛选方法": [row.get("method") for row in rows],
                    "剩余变量个数": [row.get("remaining_features") for row in rows],
                    "来源": [row.get("source") for row in rows],
                }
            )
    d03_mode = str(stage_summary.get("d03_mode", "noise_survival"))
    if d03_mode in {"feature_select_v2", "feature_select_v2_compatible", "v2"}:
        d03_label = "随机数重要性筛选（feature-select-v2兼容）：单随机列对比 split/gain 并剔除尾部重要性特征"
    else:
        d03_label = "随机噪声重要性筛选：剔除弱于噪声的真实特征"
    rows = [
        ("原始候选变量总数", "original_candidate_features"),
        ("分表基础预筛：缺失率、相关性、IV", "d01_kept_features"),
        ("稳定性筛选：DEV vs OOT PSI", "d02_kept_features"),
        ("Feather观察样本可用特征", "feather_available_features"),
        ("全局相关性去重：按单变量AUC保留更强特征", "after_global_corr"),
        (d03_label, "after_d03_random_importance"),
        ("空标签重要性筛选：保留显著高于空标签分布的特征", "after_d04_null_importance"),
        ("最终训练特征", "final_training_features"),
    ]
    return pd.DataFrame(
        {
            "步骤": range(1, len(rows) + 1),
            "筛选方法": [label for label, _ in rows],
            "剩余变量个数": [stage_summary.get(key, "N/A") for _, key in rows],
        }
    )


def _metric_comparison_frame(frame: pd.DataFrame, *, row_col: str, metric: str, row_label: str) -> pd.DataFrame:
    cols = [row_col]
    rename = {row_col: row_label}
    for score_column in SCORE_COLUMNS:
        source_col = f"{score_column}_{metric}"
        if source_col in frame.columns:
            cols.append(source_col)
            rename[source_col] = VERSION_LABELS.get(score_column, score_column)
    display = frame[cols].copy().rename(columns=rename)
    return display


def _metric_comparison_frame_long(frame: pd.DataFrame, *, row_col: str, metric: str, row_label: str) -> pd.DataFrame:
    if frame.empty or row_col not in frame.columns or metric not in frame.columns or "score_version" not in frame.columns:
        return pd.DataFrame()
    pivot = frame.pivot_table(index=row_col, columns="score_version", values=metric, aggfunc="first").reset_index()
    pivot.columns.name = None
    ordered_cols = [row_col] + [score for score in SCORE_COLUMNS if score in pivot.columns]
    display = pivot[ordered_cols].copy().sort_values(row_col)
    return display.rename(columns={row_col: row_label, **{score: VERSION_LABELS.get(score, score) for score in SCORE_COLUMNS}})


def _metric_comparison_frame_oos_by_month(frame: pd.DataFrame, *, metric: str, row_label: str) -> pd.DataFrame:
    required = {"final_flag", "mdl_month", "score_version", metric}
    if frame.empty or not required.issubset(frame.columns):
        return pd.DataFrame()
    working = frame.copy()
    working["_period_label"] = working["final_flag"].astype(str) + " " + working["mdl_month"].astype(str)
    period_order = (
        working[["_period_label", "final_flag", "mdl_month"]]
        .drop_duplicates()
        .assign(_flag_order=lambda data: data["final_flag"].map({"DEV-OOS": 0, "OOT-OOS": 1}).fillna(99))
        .sort_values(["_flag_order", "mdl_month", "_period_label"])
    )
    pivot = working.pivot_table(index="_period_label", columns="score_version", values=metric, aggfunc="first").reset_index()
    pivot.columns.name = None
    display = period_order[["_period_label"]].merge(pivot, on="_period_label", how="left")
    ordered_cols = ["_period_label"] + [score for score in SCORE_COLUMNS if score in display.columns]
    return display[ordered_cols].rename(
        columns={"_period_label": row_label, **{score: VERSION_LABELS.get(score, score) for score in SCORE_COLUMNS}}
    )


def _sloping_display_frame(frame: pd.DataFrame) -> pd.DataFrame:
    display = frame.copy().sort_values("decile", ascending=True).reset_index(drop=True)
    required = {"decile", "n_samples", "bad"}
    if not required.issubset(display.columns):
        return pd.DataFrame(
            {
                "分组": [_sloping_group_label(row) for row in display.itertuples(index=False)],
                "占比": display.get("pct"),
                "累计发起率": display.get("cum_bad_rate"),
                "累计lift": display.get("cum_lift"),
                "剩余发起率": display.get("remaining_bad_rate"),
                "剩余lift": display.get("remaining_lift"),
            }
        )

    n_samples = pd.to_numeric(display["n_samples"], errors="coerce").fillna(0)
    bad = pd.to_numeric(display["bad"], errors="coerce").fillna(0)
    total_n = float(n_samples.sum())
    total_bad = float(bad.sum())
    total_rate = total_bad / total_n if total_n else 0.0

    cum_n = n_samples.cumsum()
    cum_bad = bad.cumsum()
    remaining_n = total_n - cum_n
    remaining_bad = total_bad - cum_bad
    cum_rate = cum_bad.divide(cum_n).where(cum_n > 0, 0)
    remaining_rate = remaining_bad.divide(remaining_n).where(remaining_n > 0, 0)

    return pd.DataFrame(
        {
            "分组": [_sloping_group_label(row) for row in display.itertuples(index=False)],
            "占比": n_samples / total_n if total_n else 0,
            "累计发起率": cum_rate,
            "累计lift": cum_rate / total_rate if total_rate else 0,
            "剩余发起率": remaining_rate,
            "剩余lift": remaining_rate / total_rate if total_rate else 0,
        }
    )


def _sloping_group_label(row: Any) -> str:
    decile = int(getattr(row, "decile"))
    lower_bound = getattr(row, "lower_bound", None)
    if lower_bound is not None and not pd.isna(lower_bound):
        return f"{decile:03d}:{lower_bound}"
    return f"{decile:03d}"


def _intent_version_matrix(frame: pd.DataFrame, value_col: str) -> pd.DataFrame:
    if frame.empty or value_col not in frame.columns:
        return pd.DataFrame()
    pivot = frame.pivot_table(index="intent_level", columns="zc_level", values=value_col, aggfunc="first").reset_index()
    pivot.columns.name = None
    pivot = pivot.rename(columns={"intent_level": "意愿"})
    pivot = _sort_intent_matrix(pivot)
    preferred_cols: list[Any] = ["意愿"] + [str(value) for value in range(1, 8)] + ["合计"]
    existing_cols = [col for col in preferred_cols if col in pivot.columns]
    remaining_cols = [col for col in pivot.columns if col not in existing_cols]
    return pivot[existing_cols + remaining_cols]


def _intent_sum_matrix(frame: pd.DataFrame, value_col: str) -> pd.DataFrame:
    pivot = frame.pivot_table(index="intent_level", columns="zc_level", values=value_col, aggfunc="first").reset_index()
    pivot.columns.name = None
    pivot = pivot.rename(columns={"intent_level": "意愿"})
    pivot = _sort_intent_matrix(pivot)
    value_cols = [col for col in pivot.columns if col != "意愿"]
    if value_cols:
        pivot["sum"] = pivot[value_cols].sum(axis=1)
        total = {"意愿": "sum"}
        total.update({col: pivot[col].sum() for col in value_cols})
        total["sum"] = pivot["sum"].sum()
        pivot = pd.concat([pivot, pd.DataFrame([total])], ignore_index=True)
    return pivot


def _intent_rate_matrix(frame: pd.DataFrame, numerator_col: str, denominator_col: str) -> pd.DataFrame:
    rows = []
    work = frame.copy()
    work[numerator_col] = pd.to_numeric(work[numerator_col], errors="coerce").fillna(0)
    work[denominator_col] = pd.to_numeric(work[denominator_col], errors="coerce").fillna(0)
    zc_values = sorted(work["zc_level"].dropna().unique().tolist(), key=lambda value: str(value))
    for intent in ["低意愿", "中意愿", "高意愿"]:
        sub = work[work["intent_level"] == intent]
        if sub.empty:
            continue
        row: dict[str, Any] = {"意愿": intent}
        for zc_value in zc_values:
            cell = sub[sub["zc_level"] == zc_value]
            numerator = float(cell[numerator_col].sum())
            denominator = float(cell[denominator_col].sum())
            row[zc_value] = numerator / denominator if denominator else None
        numerator = float(sub[numerator_col].sum())
        denominator = float(sub[denominator_col].sum())
        row["sum"] = numerator / denominator if denominator else None
        rows.append(row)
    if not rows:
        return pd.DataFrame()

    total: dict[str, Any] = {"意愿": "sum"}
    for zc_value in zc_values:
        cell = work[work["zc_level"] == zc_value]
        numerator = float(cell[numerator_col].sum())
        denominator = float(cell[denominator_col].sum())
        total[zc_value] = numerator / denominator if denominator else None
    total["sum"] = float(work[numerator_col].sum()) / float(work[denominator_col].sum()) if float(work[denominator_col].sum()) else None
    rows.append(total)
    return pd.DataFrame(rows)


def _sort_intent_matrix(frame: pd.DataFrame) -> pd.DataFrame:
    if "意愿" not in frame.columns:
        return frame
    order = {"低意愿": 0, "中意愿": 1, "高意愿": 2, "sum": 3, "合计": 3}
    sorted_frame = frame.copy()
    sorted_frame["_sort"] = sorted_frame["意愿"].map(order).fillna(99)
    sorted_frame = sorted_frame.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    return sorted_frame


def _stability_pivot(frame: pd.DataFrame, value_col: str, value_label: str) -> pd.DataFrame:
    if frame.empty or value_col not in frame.columns:
        return pd.DataFrame()
    work = frame.copy()
    work["分组"] = work.apply(
        lambda row: f"{int(row['score_decile']):03d}:{row.get('lower_bound', '')}"
        if pd.notna(row.get("lower_bound", None))
        else f"{int(row['score_decile']):03d}",
        axis=1,
    )
    pivot = work.pivot_table(index=["score_decile", "分组"], columns="mdl_month", values=value_col, aggfunc="first").reset_index()
    pivot.columns.name = None
    pivot = pivot.sort_values("score_decile").drop(columns=["score_decile"])
    pivot.insert(0, "指标", value_label)
    return pivot


def _ordered_values(values: list[Any]) -> list[Any]:
    preferred = ["DEV", "DEV-OOS", "OOT", "OOT-OOS"]
    ordered = [value for value in preferred if value in values]
    ordered.extend([value for value in values if value not in ordered])
    return ordered


def _read_feature_screening_process(feature_dir: Path, *, allow_legacy_fallback: bool = False) -> dict[str, Any]:
    path = feature_dir / "feature_screening_process.json"
    if path.exists():
        return _read_json(path)
    if allow_legacy_fallback:
        legacy_path = feature_dir.parent.parent.parent / "reports" / "feature_screening_process.json"
        if legacy_path.exists():
            return _read_json(legacy_path)
    return {}


def _legacy_screening_process_allowed(feature_dir: Path) -> bool:
    try:
        run_state = load_yaml(feature_dir.parent / "run_state.yml")
    except (OSError, ValueError):
        return False
    workflow = str(run_state.get("workflow", ""))
    status = str(run_state.get("status", ""))
    return workflow.startswith("imported") or status == "imported"


def _training_feature_preparation_frame(train_dir: Path, run_config: dict[str, Any] | None = None) -> pd.DataFrame:
    run_config = run_config or _read_json(train_dir / "run_config.json")
    preprocessing = _read_json(train_dir / "preprocessing.json")
    candidate_count = _count_lines(train_dir / "candidate_feature_list.txt") or run_config.get("candidate_feature_count")
    actual_count = _count_lines(train_dir / "actual_feature_list.txt") or run_config.get("actual_feature_count")
    dropped_count = preprocessing.get("dropped_feature_count")
    if dropped_count is None and candidate_count is not None and actual_count is not None:
        try:
            dropped_count = int(candidate_count) - int(actual_count)
        except (TypeError, ValueError):
            dropped_count = "N/A"
    rows = [
        {
            "步骤": "训练输入",
            "处理说明": "读取本轮训练候选特征列表；当前 run 未登记独立特征初筛或特征精筛过程产物",
            "变量个数": candidate_count if candidate_count is not None else "N/A",
            "来源": "modeling/main_lgbm/candidate_feature_list.txt",
        },
        {
            "步骤": "训练预处理",
            "处理说明": (
                "按训练数据字段可用性、缺失哨兵、缺失率与常量字段规则保留变量；"
                f"填充策略：{preprocessing.get('fill_strategy', 'N/A')}"
            ),
            "变量个数": actual_count if actual_count is not None else "N/A",
            "来源": "modeling/main_lgbm/preprocessing.json",
        },
        {
            "步骤": "训练剔除",
            "处理说明": "训练预处理阶段剔除变量数",
            "变量个数": dropped_count if dropped_count is not None else "N/A",
            "来源": "modeling/main_lgbm/feature_drop_detail.csv",
        },
        {
            "步骤": "最终入模",
            "处理说明": "LightGBM 实际入模变量数",
            "变量个数": actual_count if actual_count is not None else "N/A",
            "来源": "modeling/main_lgbm/actual_feature_list.txt",
        },
    ]
    return pd.DataFrame(rows)


def _clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    cleaned = frame.copy()
    cleaned = cleaned.where(pd.notna(cleaned), None)
    return cleaned


def _excel_value(value: Any, column_name: str) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, str):
        return value
    if _is_count_column(column_name):
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if isinstance(value, float):
        return float(value)
    return value


def _number_format(column_name: str, value: Any) -> str:
    if value is None:
        return "General"
    if _is_count_column(column_name):
        return "#,##0"
    lowered = column_name.lower()
    name = str(column_name)
    if any(token in lowered for token in ["amount", "principal", "overdue", "amt"]) or any(token in name for token in ["金额", "本金", "逾期"]):
        return "#,##0.000"
    if _is_percent_column(column_name):
        return "0.0%"
    if isinstance(value, float):
        return "0.000"
    return "General"


def _alignment(column_name: str):
    from openpyxl.styles import Alignment

    if _is_count_column(column_name) or isinstance(column_name, str) and _looks_metric_column(column_name):
        return Alignment(horizontal="right", vertical="center", wrap_text=True)
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _is_count_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    name = str(column_name)
    if any(token in name for token in ["样本数", "样本量", "样本总量", "发起数", "发起量", "正样本", "风险数"]):
        return True
    return any(token in lowered for token in ["count", "samples", "positive", "bad", "split", "unique"]) and not any(
        token in lowered for token in ["rate", "ratio", "pct", "lift", "auc", "ks"]
    )


def _is_percent_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    name = str(column_name)
    if any(token in lowered for token in ["auc", "ks", "lift"]):
        return False
    return any(token in lowered for token in ["rate", "ratio", "pct"]) or any(token in name for token in ["率", "占比"])


def _looks_metric_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    return any(
        token in lowered
        for token in [
            "auc",
            "ks",
            "rate",
            "ratio",
            "pct",
            "lift",
            "psi",
            "gain",
            "score",
            "amount",
            "principal",
            "overdue",
            "value",
            "发起率",
            "占比",
            "金额逾期率",
            "人头风险率",
        ]
    )


def _style_region(ws, start_row: int, end_row: int, ncols: int, *, start_col: int = 1) -> None:
    from openpyxl.styles import Border, Font, Side

    thin = Side(style="thin", color="D9E0E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=start_col + ncols - 1):
        for cell in row:
            cell.border = border
            if cell.font is None or cell.font.name is None:
                cell.font = Font(name="楷体", size=10)


def _apply_table_color_scales(ws, header_row: int, end_row: int, frame: pd.DataFrame, *, start_col: int = 1) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    if end_row <= header_row:
        return

    for col_offset, column_name in enumerate(frame.columns):
        lowered = str(column_name).lower()
        if not _should_color_scale(lowered):
            continue
        col_letter = get_column_letter(start_col + col_offset)
        cell_range = f"{col_letter}{header_row + 1}:{col_letter}{end_row}"
        if any(token in lowered for token in ["uplift", "提升"]):
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFFFF",
                end_type="max",
                end_color="63BE7B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            )
        ws.conditional_formatting.add(cell_range, rule)


def _apply_color_scale(ws, start_row: int, end_row: int, start_col: int, end_col: int, *, prefer_high: bool) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    if end_row < start_row or end_col < start_col:
        return
    if prefer_high:
        start_color, end_color = "63BE7B", "F8696B"
    else:
        start_color, end_color = "F8696B", "63BE7B"
    rule = ColorScaleRule(
        start_type="min",
        start_color=start_color,
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color=end_color,
    )
    ws.conditional_formatting.add(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}", rule)


def _should_color_scale(lowered_column_name: str) -> bool:
    return any(
        token in lowered_column_name
        for token in [
            "uplift",
            "cum_lift",
            "remaining_lift",
            "bad_rate",
            "head_risk_rate",
            "amount_overdue_rate",
            "psi",
            "占比",
            "发起率",
            "金额逾期率",
            "人头风险率",
        ]
    )


def _update_column_widths(ws, frame: pd.DataFrame, ncols: int, *, start_col: int = 1) -> None:
    from openpyxl.utils import get_column_letter

    for col_offset in range(ncols):
        col_name = str(frame.columns[col_offset])
        max_len = len(col_name)
        sample_values = frame.iloc[:80, col_offset].tolist()
        for value in sample_values:
            if value is not None and not pd.isna(value):
                max_len = max(max_len, len(str(value)))
        if any(token in col_name.lower() for token in ["feature", "desc", "data_source", "内容"]) or any(
            token in col_name for token in ["结论", "摘要"]
        ):
            width = min(max(max_len * 1.1, 18), 52)
        elif _looks_metric_column(col_name) or _is_count_column(col_name):
            width = min(max(max_len * 0.85, 10), 16)
        else:
            width = min(max(max_len * 1.0, 10), 24)
        ws.column_dimensions[get_column_letter(start_col + col_offset)].width = width


def _finalize_sheet(ws) -> None:
    from copy import copy

    from openpyxl.styles import Alignment, Font

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_format.defaultRowHeight = 18
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                font = copy(cell.font)
                if font.name is None:
                    font.name = "楷体"
                    cell.font = font
                alignment = copy(cell.alignment) if cell.alignment else Alignment()
                if alignment.vertical is None:
                    alignment.vertical = "center"
                if alignment.wrap_text is None:
                    alignment.wrap_text = True
                cell.alignment = alignment
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = "E7F4F2"
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws["A1"] = "暂无可用数据"
        ws["A1"].font = Font(name="楷体", size=11, italic=True)


def _write_model_reports(
    *,
    output_path: Path,
    train_dir: Path,
    eval_dir: Path,
    feature_dir: Path,
    sample_dir: Path,
    include_gcard_summary: bool = False,
) -> tuple[Path, Path]:
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    stage_summary = _read_json(feature_dir / "feature_stage_summary.json")
    screening_process = _read_feature_screening_process(feature_dir, allow_legacy_fallback=_legacy_screening_process_allowed(feature_dir))
    overall = _read_csv(eval_dir / "overall_metrics.csv")
    benchmark = _read_csv(eval_dir / "benchmark_uplift.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    versioned_intent = _read_csv(eval_dir / "intent_zc_segment_distribution_by_version.csv")
    model_score_stability = _read_csv(eval_dir / "model_score_bin_distribution_by_month.csv")
    importance = _read_csv(train_dir / "feature_importance.csv")
    sample_split = _read_csv(sample_dir / "sample_split_summary.csv")
    woe_summary_path = _find_woe_summary(train_dir=train_dir, report_dir=output_path.parent)
    woe_summary = _read_csv(woe_summary_path) if woe_summary_path else None

    md_path = output_path.with_name("model_report.md")
    html_path = output_path.with_name("model_report.html")

    final_features = stage_summary.get("final_training_features", run_config.get("actual_feature_count", "N/A"))
    valid_auc = _fmt_metric(metrics.get("valid_auc"))
    valid_ks = _fmt_metric(metrics.get("valid_ks"))
    auc_gap = _fmt_metric(metrics.get("auc_gap"))
    oot_oos = _row_by_value(benchmark, "final_flag", "OOT-OOS")

    lines = [
        f"# {REPORT_TITLE}",
        "",
        f"生成日期：{time.strftime('%Y-%m-%d')}",
        "",
    ]
    if include_gcard_summary:
        lines.extend(_gcard_summary_markdown_lines(train_dir=train_dir, eval_dir=eval_dir, feature_dir=feature_dir))
        lines.append("")
    lines.extend(
        [
        "## 一、模型描述",
        "",
        f"- 模型目标：预测配置标签字段 `{run_config.get('label_column', 'target')}`。",
        f"- 建模样本：训练集 {_fmt_list(run_config.get('train_values', []))}，验证集 {_fmt_list(run_config.get('valid_values', []))}，OOS {_fmt_list(run_config.get('oos_values', []))}。",
        f"- 算法：{run_config.get('algorithm', 'N/A')}；最终入模变量 {final_features} 个；best iteration {run_config.get('best_iteration', 'N/A')}。",
        f"- 验证集效果：AUC {valid_auc}，KS {valid_ks}，Train/Valid AUC gap {auc_gap}。",
        "",
        ]
    )
    lines.extend(["## 二、变量筛选过程", ""])
    if screening_process.get("feature_select_v2_alignment", {}).get("summary"):
        lines.append(f"- {screening_process['feature_select_v2_alignment']['summary']}")
        lines.append("")
    if screening_process.get("screening_rows"):
        screening_frame = pd.DataFrame(screening_process["screening_rows"]).rename(
            columns={"step": "步骤", "method": "筛选方法", "remaining_features": "剩余变量个数", "source": "来源"}
        )
    elif stage_summary:
        screening_frame = _screening_steps_frame(stage_summary, feature_dir)
    else:
        lines.append("- 当前 run 未登记独立特征初筛或特征精筛过程产物；以下只展示训练阶段实际候选和入模特征准备结果。")
        lines.append("")
        screening_frame = _training_feature_preparation_frame(train_dir, run_config)
    lines.extend(_markdown_table(screening_frame))
    lines.append("")

    preprocessing = _read_json(train_dir / "preprocessing.json")
    leakage_warnings = _feature_list_warnings(train_dir)
    if preprocessing:
        lines.append(
            f"- 训练预处理保留 {preprocessing.get('kept_feature_count', run_config.get('actual_feature_count', 'N/A'))}/"
            f"{preprocessing.get('candidate_feature_count', run_config.get('candidate_feature_count', 'N/A'))} 个变量；"
            f"剔除 {preprocessing.get('dropped_feature_count', 'N/A')} 个变量。"
        )
    if leakage_warnings:
        lines.append("- 特征列表包含需复核提示：" + "；".join(leakage_warnings[:3]))
    lines.append("")
    lines.extend(
        [
            "",
            "## 三、核心效果与历史版本对比",
            "",
            _metric_sentence("OOT-OOS", oot_oos),
            "",
        ]
    )
    if overall is not None:
        lines.extend(_markdown_table(overall[["final_flag", "n_samples", "positive", "bad_rate", "model_score_auc", "model_score_ks"]]))
        lines.append("")
    if benchmark is not None:
        display_cols = ["final_flag", "model_score_auc", "model_score_ks"]
        display_cols.extend(f"ks_uplift_vs_{score}" for score in _comparison_score_columns())
        lines.extend(_markdown_table(benchmark[[col for col in display_cols if col in benchmark.columns]]))
        lines.append("")

    lines.extend(["## 四、模型效果", ""])
    _append_monthly_effect_markdown(lines, eval_dir)
    _append_sloping_markdown(lines, eval_dir)
    _append_intent_risk_markdown(lines, eval_dir)

    lines.extend(["## 五、模型稳定性", ""])
    if psi is not None and not psi.empty:
        if model_score_stability is not None and not model_score_stability.empty:
            max_psi = (
                model_score_stability[["mdl_month", "month_psi"]]
                .drop_duplicates()
                .sort_values("month_psi", ascending=False)
                .head(5)
            )
            lines.append("- 本轮模型稳定性已补齐分箱占比、分箱发起率和 PSI 组件；月度 PSI 最高的 5 个观测如下：")
            lines.extend(_markdown_table(max_psi))
        else:
            if "score_column" in psi.columns:
                psi = psi[psi["score_column"] == "model_score"].copy()
            max_psi = psi.sort_values("psi", ascending=False).head(5)
            lines.append("- 本轮模型 PSI 最高的 5 个观测如下：")
            lines.extend(_markdown_table(max_psi))
    lines.extend(
        [
            "",
            "## 六、重要变量",
            "",
        ]
    )
    if importance is not None:
        lines.extend(_markdown_table(importance.head(15)))
        lines.append("")
    lines.extend(
        [
            "## 七、Top变量WOE",
            "",
        ]
    )
    if woe_summary is not None and not woe_summary.empty and "status" in woe_summary.columns and (woe_summary["status"] == "ok").any():
        lines.append("- Top20 WOE 图见 Excel sheet `Top变量WOE`，PNG 和汇总 CSV 见 `reports/woe_top_features/` 或训练产物目录。")
        display = (
            woe_summary[woe_summary["status"] == "ok"]
            .groupby(["rank", "feature"], as_index=False)
            .agg({"gain": "first", "iv_component": "sum"})
            .rename(columns={"rank": "排名", "feature": "变量", "gain": "Gain", "iv_component": "IV"})
            .sort_values("排名")
        )
        lines.extend(_markdown_table(display, limit=20))
        lines.append("")
    else:
        lines.append("- 暂无 Top20 WOE 图；该产物需要训练阶段保留 row-level 特征值后生成。")
        lines.append("")
    lines.extend(
        [
            "## 八、待补充事项",
            "",
            "- 当前仍不可补齐：变量分布/分箱图、变量中文描述与业务标签、MOB1/MOB3 历史风险精确定义；这些需要原始特征值、业务字典或未来期还款表现数据。",
            "- 详见 `model_report_missing_results.md`。",
        ]
    )

    markdown = "\n".join(lines).rstrip() + "\n"
    md_path.write_text(markdown, encoding="utf-8")
    html_path.write_text(_markdown_to_simple_html(markdown), encoding="utf-8")
    return md_path, html_path


def _append_module_conclusions_markdown(lines: list[str], eval_dir: Path, module: str) -> None:
    frame = _module_conclusion_frame(eval_dir, module)
    if frame.empty:
        return
    lines.extend(_markdown_table(frame, limit=50))
    lines.append("")


def _append_monthly_effect_markdown(lines: list[str], eval_dir: Path) -> None:
    monthly_segment_oos = _read_csv(eval_dir / "monthly_segment_metrics_oos_by_version.csv")
    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    feb_apr = _read_csv(eval_dir / "feb_apr_2026_monthly_metrics.csv")
    feb_apr_segment = _read_csv(eval_dir / "feb_apr_2026_segment_monthly.csv")
    if (
        (monthly_segment_oos is None or monthly_segment_oos.empty)
        and (monthly is None or monthly.empty)
        and (segment is None or segment.empty)
        and (feb_apr is None or feb_apr.empty)
        and (feb_apr_segment is None or feb_apr_segment.empty)
    ):
        return
    lines.extend(["1、每月效果（OOS）", ""])
    _append_module_conclusions_markdown(lines, eval_dir, "1、每月效果（OOS）")

    if monthly_segment_oos is not None and not monthly_segment_oos.empty:
        for segment_name in ["全客群", "老户次新", "老户", "次新", "流失户"]:
            subset = monthly_segment_oos[
                (monthly_segment_oos["segment"] == segment_name)
                & (monthly_segment_oos["final_flag"].isin(["DEV-OOS", "OOT-OOS"]))
            ].copy()
            if subset.empty:
                continue
            lines.append(f"在{segment_name} OOS by月效果（KS）")
            lines.extend(_markdown_table(_metric_comparison_frame_oos_by_month(subset, metric="ks", row_label="样本月份"), limit=50))
            lines.append("")
            lines.append(f"在{segment_name} OOS by月效果（AUC）")
            lines.extend(_markdown_table(_metric_comparison_frame_oos_by_month(subset, metric="auc", row_label="样本月份"), limit=50))
            lines.append("")
    elif monthly is not None and not monthly.empty:
        working = monthly.copy()
        working["_period_label"] = working["final_flag"].astype(str) + " " + working["mdl_month"].astype(str)
        lines.append("全客群 by月效果（KS）")
        lines.extend(_markdown_table(_metric_comparison_frame(working, row_col="_period_label", metric="ks", row_label="样本月份"), limit=80))
        lines.append("")
        lines.append("全客群 by月效果（AUC）")
        lines.extend(_markdown_table(_metric_comparison_frame(working, row_col="_period_label", metric="auc", row_label="样本月份"), limit=80))
        lines.append("")

    if segment is not None and not segment.empty:
        lines.append("分客群整体效果（KS）")
        segment_ks_rows = []
        for segment_name in ["全客群", "老户次新", "老户", "次新", "流失户"]:
            subset = segment[segment["segment"] == segment_name].copy()
            if subset.empty:
                continue
            table = _metric_comparison_frame(subset, row_col="final_flag", metric="ks", row_label="样本")
            if table.empty:
                continue
            table.insert(0, "客群", segment_name)
            segment_ks_rows.extend(table.to_dict("records"))
        if segment_ks_rows:
            lines.extend(_markdown_table(pd.DataFrame(segment_ks_rows), limit=80))
            lines.append("")
        lines.append("分客群整体效果（AUC）")
        segment_auc_rows = []
        for segment_name in ["全客群", "老户次新", "老户", "次新", "流失户"]:
            subset = segment[segment["segment"] == segment_name].copy()
            if subset.empty:
                continue
            table = _metric_comparison_frame(subset, row_col="final_flag", metric="auc", row_label="样本")
            if table.empty:
                continue
            table.insert(0, "客群", segment_name)
            segment_auc_rows.extend(table.to_dict("records"))
        if segment_auc_rows:
            lines.extend(_markdown_table(pd.DataFrame(segment_auc_rows), limit=80))
            lines.append("")

    if feb_apr is not None and not feb_apr.empty:
        lines.append("2026年2-4月外推验证（全客群）")
        display_cols = ["month", "split", "n_samples", "bad_rate", "model_auc", "model_ks", "v6_auc", "v6_ks"]
        lines.extend(_markdown_table(feb_apr[[col for col in display_cols if col in feb_apr.columns]], limit=80))
        lines.append("")
    if feb_apr_segment is not None and not feb_apr_segment.empty:
        lines.append("2026年2-4月外推验证（分客群）")
        display_cols = ["month", "split", "segment", "n_samples", "bad_rate", "model_auc", "model_ks", "v6_auc", "v6_ks", "ks_uplift"]
        lines.extend(_markdown_table(feb_apr_segment[[col for col in display_cols if col in feb_apr_segment.columns]], limit=80))
        lines.append("")


def _append_sloping_markdown(lines: list[str], eval_dir: Path) -> None:
    versioned = _read_csv(eval_dir / "decile_lift_bins_by_version.csv")
    if versioned is None or versioned.empty:
        available = [
            (segment_name, segment_key, score_column)
            for segment_name, segment_key in SEGMENT_FILES.items()
            for score_column in SCORE_COLUMNS
            if (eval_dir / f"decile_lift_{segment_key}_{score_column}.csv").exists()
            or (score_column == "model_score" and (eval_dir / f"decile_lift_{segment_key}.csv").exists())
        ]
        if not available:
            return
        lines.extend(["2、模型sloping", ""])
        _append_module_conclusions_markdown(lines, eval_dir, "2、模型sloping")
        for segment_name, segment_key, score_column in available:
            frame = _read_csv(eval_dir / f"decile_lift_{segment_key}_{score_column}.csv")
            if frame is None and score_column == "model_score":
                frame = _read_csv(eval_dir / f"decile_lift_{segment_key}.csv")
            if frame is None or frame.empty:
                continue
            lines.append(f"全样本 30天发起：在{segment_name}效果 - {VERSION_LABELS.get(score_column, score_column)}")
            lines.extend(_markdown_table(_sloping_display_frame(frame), limit=20))
            lines.append("")
        return

    lines.extend(["2、模型sloping", ""])
    _append_module_conclusions_markdown(lines, eval_dir, "2、模型sloping")
    for segment_name in ["全客群", "老户次新", "流失户"]:
        subset_segment = versioned[(versioned["segment"] == segment_name) & (versioned["final_flag"] == "OOT-OOS")].copy()
        if subset_segment.empty:
            continue
        lines.append(f"OOT-OOS 30天发起：在{segment_name}效果")
        lines.append("")
        for score_column in SCORE_COLUMNS:
            subset = subset_segment[subset_segment["score_version"] == score_column].copy()
            if subset.empty:
                continue
            lines.append(VERSION_LABELS.get(score_column, score_column))
            lines.extend(_markdown_table(_sloping_display_frame(subset), limit=20))
            lines.append("")


def _append_intent_risk_markdown(lines: list[str], eval_dir: Path) -> None:
    distribution = _read_csv(eval_dir / "intent_zc_segment_distribution_by_version.csv")
    ftr_rate = _read_csv(eval_dir / "intent_zc_segment_ftr_rate_by_version.csv")
    amount_risk = _read_csv(eval_dir / "intent_zc_segment_amount_risk_by_version.csv")
    if (
        distribution is None
        or distribution.empty
        or ftr_rate is None
        or ftr_rate.empty
        or amount_risk is None
        or amount_risk.empty
    ):
        basic_distribution = _read_csv(eval_dir / "intent_zc_distribution.csv")
        basic_amount_risk = _read_csv(eval_dir / "intent_zc_amount_risk.csv")
        basic_head_risk = _read_csv(eval_dir / "intent_zc_headcount_risk.csv")
        if (
            (basic_distribution is None or basic_distribution.empty)
            and (basic_amount_risk is None or basic_amount_risk.empty)
            and (basic_head_risk is None or basic_head_risk.empty)
        ):
            return
        lines.extend(["3、意愿交叉风险（DEV-OOS）", ""])
        lines.append("- 当前 run 仅有全量观察口径的意愿 x 资产评级结果，缺少老户/流失户和历史版本分层矩阵；完整缺失项见 `model_report_missing_results.md`。")
        lines.append("")
        if basic_distribution is not None and not basic_distribution.empty:
            lines.append("当前可用全量观察：占比（意愿评级 x 资产评级）")
            lines.extend(_markdown_table(_intent_sum_matrix(basic_distribution, "pct"), limit=20))
            lines.append("")
            lines.append("当前可用全量观察：30天发起率（意愿评级 x 资产评级）")
            lines.extend(_markdown_table(_intent_rate_matrix(basic_distribution, "bad", "n_samples"), limit=20))
            lines.append("")
        if basic_head_risk is not None and not basic_head_risk.empty:
            lines.append("当前可用全量观察：人头风险率（意愿评级 x 资产评级）")
            lines.extend(_markdown_table(_intent_rate_matrix(basic_head_risk, "head_risk_count", "n_samples"), limit=20))
            lines.append("")
        if basic_amount_risk is not None and not basic_amount_risk.empty:
            lines.append("当前可用全量观察：金额风险（仅意愿维度）")
            lines.extend(_markdown_table(basic_amount_risk, limit=20))
            lines.append("")
        return
    lines.extend(["3、意愿交叉风险（DEV-OOS）", ""])
    _append_module_conclusions_markdown(lines, eval_dir, "3、意愿交叉风险（DEV-OOS）")
    matrix_specs = [
        ("占比", distribution, "sample_pct"),
        ("30天发起率", ftr_rate, "ftr_30d_rate"),
        ("新增订单3期金额逾期率", amount_risk, "amount_overdue_rate"),
    ]
    for segment_name in ["老户", "流失户"]:
        lines.append(f"{segment_name} DEV-OOS 意愿 x 资产评级")
        lines.append("")
        for metric_label, frame, value_col in matrix_specs:
            metric_frame = frame[(frame["segment"] == segment_name) & (frame["final_flag"] == "DEV-OOS")].copy()
            if metric_frame.empty:
                continue
            lines.append(metric_label)
            lines.append("")
            for score_column in SCORE_COLUMNS:
                subset = metric_frame[metric_frame["score_version"] == score_column].copy()
                if subset.empty:
                    continue
                lines.append(f"{segment_name} - {metric_label} - {VERSION_LABELS.get(score_column, score_column)}")
                lines.extend(_markdown_table(_intent_version_matrix(subset, value_col), limit=20))
                lines.append("")


def _fmt_list(values: Any) -> str:
    if isinstance(values, list):
        return "、".join(str(value) for value in values) if values else "N/A"
    return str(values)


def _metric_sentence(label: str, row: dict[str, Any]) -> str:
    if not row:
        return f"- {label} 暂无可用 benchmark 指标。"
    comparisons = []
    for score_column in _comparison_score_columns():
        uplift = row.get(f"ks_uplift_vs_{score_column}")
        if uplift is not None:
            comparisons.append(f"相对 {VERSION_LABELS.get(score_column, score_column)} KS 提升 {_fmt_metric(uplift)}")
    comparison_text = "，".join(comparisons) if comparisons else "未配置可比 champion 分数"
    return (
        f"- {label}：本轮模型 AUC {_fmt_metric(row.get('model_score_auc'))}、KS {_fmt_metric(row.get('model_score_ks'))}；"
        f"{comparison_text}。"
    )


def _row_by_value(frame: pd.DataFrame | None, column: str, value: Any) -> dict[str, Any]:
    if frame is None or frame.empty or column not in frame.columns:
        return {}
    matched = frame[frame[column] == value]
    if matched.empty:
        return {}
    return matched.iloc[0].to_dict()


def _markdown_table(frame: pd.DataFrame, *, limit: int = 20) -> list[str]:
    if frame.empty:
        return ["暂无可用数据"]
    display = frame.head(limit).copy()
    display = display.rename(columns={col: str(col) for col in display.columns})
    rows = []
    headers = [str(col) for col in display.columns]
    rows.append("| " + " | ".join(headers) + " |")
    rows.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for item in display.itertuples(index=False):
        rows.append("| " + " | ".join(_fmt_markdown_cell(value) for value in item) + " |")
    if len(frame) > limit:
        rows.append(f"\n> 仅展示前 {limit} 行，完整明细见 Excel。")
    return rows


def _fmt_markdown_cell(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.3f}"
    return str(value).replace("|", "\\|")


def _fmt_metric(value: Any) -> str:
    if value is None or value == "N/A":
        return "N/A"
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)


def _fmt_pp(value: Any) -> str:
    try:
        return f"{float(value) * 100:.1f}"
    except (TypeError, ValueError):
        return "N/A"


def _fmt_percent(value: Any) -> str:
    try:
        return f"{float(value):.1%}"
    except (TypeError, ValueError):
        return "N/A"


def _markdown_to_simple_html(markdown: str) -> str:
    lines = markdown.splitlines()
    html_lines = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head>",
        '<meta charset="utf-8">',
        f"<title>{escape(REPORT_TITLE)}</title>",
        "<style>",
        "body{font-family:Arial,'Microsoft YaHei',sans-serif;margin:32px;color:#1f2933;line-height:1.55}",
        "h1{font-size:26px;margin-bottom:16px} h2{font-size:19px;margin-top:28px;border-bottom:1px solid #d9e0e3;padding-bottom:6px}",
        "table{border-collapse:collapse;margin:12px 0 20px 0;font-size:13px} th,td{border:1px solid #d9e0e3;padding:6px 8px;text-align:right} th{background:#f3f6f6} td:first-child,th:first-child{text-align:left}",
        "li{margin:4px 0} blockquote{color:#667085;border-left:3px solid #d9e0e3;padding-left:10px}",
        "</style>",
        "</head><body>",
    ]
    in_ul = False
    in_table = False
    for line in lines:
        if line.startswith("# "):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            html_lines.append(f"<h1>{escape(line[2:])}</h1>")
        elif line.startswith("## "):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            html_lines.append(f"<h2>{escape(line[3:])}</h2>")
        elif line.startswith("- "):
            if not in_ul:
                html_lines.append("<ul>")
                in_ul = True
            html_lines.append(f"<li>{escape(line[2:])}</li>")
        elif line.startswith("| ") and line.endswith(" |"):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if all(cell == "---" for cell in cells):
                continue
            if not in_table:
                html_lines.append("<table>")
                in_table = True
                tag = "th"
            else:
                tag = "td"
            html_lines.append("<tr>" + "".join(f"<{tag}>{escape(cell)}</{tag}>" for cell in cells) + "</tr>")
        elif line.startswith("> "):
            if in_table:
                html_lines.append("</table>")
                in_table = False
            html_lines.append(f"<blockquote>{escape(line[2:])}</blockquote>")
        else:
            if in_table:
                html_lines.append("</table>")
                in_table = False
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            if line.strip():
                html_lines.append(f"<p>{escape(line)}</p>")
    if in_table:
        html_lines.append("</table>")
    if in_ul:
        html_lines.append("</ul>")
    html_lines.append("</body></html>")
    return "\n".join(html_lines)


def _write_missing_results_doc(output_path: Path, *, train_dir: Path | None = None, eval_dir: Path | None = None) -> Path:
    missing_path = output_path.with_name("model_report_missing_results.md")
    eval_dir = eval_dir or output_path.parent.parent / "evaluation"
    comparison_scores = _comparison_score_columns()
    historical_section = ""
    historical_rows = [
        ("intent_zc_segment_distribution_by_version.csv", "各版本老户/流失户 DEV-OOS 意愿资产占比矩阵"),
        ("intent_zc_segment_ftr_rate_by_version.csv", "各版本老户/流失户 DEV-OOS 30天发起率矩阵"),
        ("intent_zc_segment_amount_risk_by_version.csv", "各版本老户/流失户 DEV-OOS 新增订单3期金额逾期率矩阵"),
        ("decile_lift_bins_by_version.csv", "各版本 sloping 分箱上下界"),
        ("monthly_segment_metrics_oos_by_version.csv", "全客群/老户次新/老户/次新/流失户 DEV-OOS + OOT-OOS 按月版本横向效果"),
        ("monthly_segment_metrics_oot_oos_by_version.csv", "老户次新/流失户 OOT-OOS 按月版本横向效果"),
        ("score_bin_distribution_by_month_by_version.csv", "各版本按月稳定性分箱"),
    ]
    existing_historical_rows = [(name, desc) for name, desc in historical_rows if (eval_dir / name).exists()]
    if existing_historical_rows:
        score_versions = _score_versions_in_files(eval_dir, [name for name, _ in existing_historical_rows])
        if not score_versions:
            score_versions = ["model_score", *comparison_scores]
        covered_versions = "、".join(f"`{score}`" for score in score_versions)
        historical_table = "\n".join(f"| `evaluation/{name}` | {desc} |" for name, desc in existing_historical_rows)
        historical_section = f"""
## 已补齐 — 历史版本横向对比

覆盖 score_version：{covered_versions}。

| 产出文件 | 内容 |
|---|---|
{historical_table}
"""
    model_score_rows = [
        ("monthly_metrics.csv", "全客群按月 AUC/KS"),
        ("segment_metrics.csv", "全客群/老户次新/老户/次新/流失户切片 AUC/KS"),
        ("benchmark_uplift.csv", "本轮模型 vs 历史版本整体提升量"),
        ("score_psi_by_month.csv", "分数 PSI 稳定性"),
        ("decile_lift_all_model_score.csv", "全客群本轮模型十分位 lift"),
        ("decile_lift_e2e3_model_score.csv", "老户次新本轮模型十分位 lift"),
        ("decile_lift_b2_model_score.csv", "流失户本轮模型十分位 lift"),
        ("intent_zc_distribution.csv", "全量观察口径意愿资产占比和发起率矩阵"),
        ("intent_zc_amount_risk.csv", "全量观察口径意愿维度金额风险"),
        ("intent_zc_headcount_risk.csv", "全量观察口径意愿资产人头风险矩阵"),
        ("feb_apr_2026_monthly_metrics.csv", "2026年2-4月外推全客群指标"),
        ("feb_apr_2026_segment_monthly.csv", "2026年2-4月外推分客群指标"),
    ]
    existing_model_rows = [(name, desc) for name, desc in model_score_rows if (eval_dir / name).exists()]
    model_score_section = ""
    if existing_model_rows:
        model_score_table = "\n".join(f"| `evaluation/{name}` | {desc} |" for name, desc in existing_model_rows)
        model_score_section = f"""
## 已补齐 — 本轮 model_score

| 产出文件 | 内容 |
|---|---|
{model_score_table}
"""
    text = f"""# {REPORT_TITLE}缺失结果清单

本文件只记录当前已注册 run artifact 无法可靠还原的内容，不伪造指标。

## 不可补齐（3 项）

| # | 缺少字段/结果 | 原因 |
|---|---|---|
| 1 | 变量分布/分箱图 | 当前评分 feather 仅含模型分数和标签，不含原始特征值 |
| 2 | 变量中文描述、业务标签 | 需要业务知识、特征字典或业务标签 |
| 3 | MOB1/MOB3 历史风险 | 需要未来期还款表现数据，当前数据仅含 30 天发起标签和观察风险字段 |
{model_score_section}
{historical_section}
{_woe_missing_results_section(train_dir=train_dir, report_dir=output_path.parent)}

后续继续通过 `rmw report` 统一生成报告，避免人工改写 xlsx 中的计算结果。
"""
    missing_path.write_text(text, encoding="utf-8")
    return missing_path


def _score_versions_in_files(eval_dir: Path, file_names: list[str]) -> list[str]:
    versions: list[str] = []
    for file_name in file_names:
        frame = _read_csv(eval_dir / file_name)
        if frame is None or "score_version" not in frame.columns:
            continue
        for value in frame["score_version"].dropna().astype(str).unique().tolist():
            if value not in versions:
                versions.append(value)
    ordered = [score for score in SCORE_COLUMNS if score in versions]
    ordered.extend([score for score in versions if score not in ordered])
    return ordered


def _woe_missing_results_section(*, train_dir: Path | None, report_dir: Path) -> str:
    summary_path = _find_woe_summary(train_dir=train_dir, report_dir=report_dir) if train_dir is not None else None
    if summary_path and any((summary_path.parent / "images").glob("*_WOE.png")):
        return """
## 已补齐 — Top20 变量 WOE

| 产出文件 | 内容 |
|---|---|
| `reports/woe_top_features/woe_top20_summary.csv` | Top20 变量分箱、WOE、IV 和人群占比 |
| `reports/woe_top_features/images/` | Top20 变量 WOE 折线与人群占比柱图 |
"""
    if summary_path:
        return """
## 部分生成 — Top20 变量 WOE

| 产出文件 | 内容 |
|---|---|
| `reports/woe_top_features/woe_top20_summary.csv` | Top20 变量分箱、WOE、IV 和人群占比 |

WOE 汇总表已生成，但 PNG 图未生成或未注册。通常原因是当前环境缺少 `matplotlib`。
"""
    return """
## 待生成 — Top20 变量 WOE

Top20 变量 WOE 图需要训练阶段保留 row-level 特征值。当前已注册 artifact 不包含原始特征值，因此报告只写缺失说明，不伪造 WOE 图。
"""


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv(path: Path) -> pd.DataFrame | None:
    if not path.exists():
        return None
    return pd.read_csv(path, encoding="utf-8-sig")


def _first_existing(*paths: Path) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def _feature_count(train_dir: Path, feature_dir: Path) -> int:
    return len(_read_feature_list(train_dir, feature_dir))


def _read_feature_list(train_dir: Path, feature_dir: Path) -> list[str]:
    for path in [
        feature_dir / "final_features.txt",
        feature_dir / "feature_list.txt",
        train_dir / "actual_feature_list.txt",
    ]:
        if path.exists():
            return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    return []


def _count_lines(path: Path) -> int | None:
    if not path.exists():
        return None
    return len([line for line in path.read_text(encoding="utf-8").splitlines() if line.strip()])


def _feature_list_warnings(train_dir: Path) -> list[str]:
    warnings: list[str] = []
    for path in [train_dir / "feature_list.txt", train_dir / "candidate_feature_list.txt"]:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if stripped.startswith("#") and ("WARN" in stripped.upper() or "LEAKAGE" in stripped.upper()):
                warnings.append(stripped.lstrip("#").strip())
        if warnings:
            break
    return warnings
