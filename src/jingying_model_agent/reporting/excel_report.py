"""Excel report generation from registered modeling artifacts."""

from __future__ import annotations

import json
import time
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd


REPORT_SHEETS = [
    "模型描述",
    "重要变量",
    "变量筛选过程和模型参数",
    "模型效果-每月效果",
    "模型效果-模型sloping",
    "模型效果-意愿交叉风险（DEV-OOS）",
    "模型稳定性",
]

SCORE_COLUMNS = ["model_score", "gcard_v2", "gcard_v4", "gcard_v5", "gcard_v6"]
VERSION_LABELS = {
    "model_score": "本轮模型",
    "gcard_v2": "G卡V2",
    "gcard_v4": "G卡V4",
    "gcard_v5": "G卡V5",
    "gcard_v6": "G卡V6",
}
SEGMENT_FILES = {
    "全客群": "all",
    "老户次新": "e2e3",
    "流失户": "b2",
}


def generate_excel_report(
    *,
    eval_dir: str | Path,
    train_dir: str | Path,
    input_dir: str | Path,
    feature_dir: str | Path,
    output_path: str | Path,
) -> Path:
    """Generate a formatted Excel model report from standard artifact folders."""
    from openpyxl import Workbook

    eval_dir = Path(eval_dir)
    train_dir = Path(train_dir)
    input_dir = Path(input_dir)
    feature_dir = Path(feature_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    run_dir = output_path.parent.parent
    sample_dir = run_dir / "sample_check"

    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_SHEETS[0]
    for sheet_name in REPORT_SHEETS[1:]:
        wb.create_sheet(sheet_name)

    _build_description_sheet(
        wb["模型描述"],
        train_dir=train_dir,
        eval_dir=eval_dir,
        input_dir=input_dir,
        sample_dir=sample_dir,
        feature_dir=feature_dir,
    )
    _build_features_sheet(wb["重要变量"], train_dir=train_dir, feature_dir=feature_dir)
    _build_screening_params_sheet(wb["变量筛选过程和模型参数"], train_dir=train_dir, feature_dir=feature_dir)
    _build_monthly_effect_sheet(wb["模型效果-每月效果"], eval_dir=eval_dir)
    _build_sloping_sheet(wb["模型效果-模型sloping"], eval_dir=eval_dir)
    _build_intent_risk_sheet(wb["模型效果-意愿交叉风险（DEV-OOS）"], eval_dir=eval_dir)
    _build_stability_sheet(wb["模型稳定性"], eval_dir=eval_dir)

    for worksheet in wb.worksheets:
        _finalize_sheet(worksheet)

    wb.save(str(output_path))
    _write_missing_results_doc(output_path)
    _write_model_reports(
        output_path=output_path,
        train_dir=train_dir,
        eval_dir=eval_dir,
        feature_dir=feature_dir,
        sample_dir=sample_dir,
    )
    return output_path


def _build_description_sheet(
    ws,
    *,
    train_dir: Path,
    eval_dir: Path,
    input_dir: Path,
    sample_dir: Path,
    feature_dir: Path,
) -> None:
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    eval_summary = _read_json(eval_dir / "evaluation_summary.json")
    feature_count = _feature_count(train_dir, feature_dir)

    row = 1
    row = _write_kv_section(
        ws,
        row,
        "报告信息",
        [
            ("报告名称", "复借G卡模型报告"),
            ("生成日期", time.strftime("%Y-%m-%d")),
            ("模型口径", "本轮基于新数据、新特征集合和新时间窗口训练主模型"),
            ("评估说明", "分客群结果作为效果切片评估，是否分客群训练由建模需求文档决定"),
        ],
    )

    row = _write_kv_section(
        ws,
        row,
        "样本与训练配置",
        [
            ("标签字段", run_config.get("label_column", "ftr_30d_ord_flag")),
            ("切分字段", run_config.get("split_column", "final_flag")),
            ("训练样本", f"final_flag in {run_config.get('train_values', [])}"),
            ("验证样本", f"final_flag in {run_config.get('valid_values', [])}"),
            ("OOS 样本", f"final_flag in {run_config.get('oos_values', [])}"),
            ("算法", run_config.get("algorithm", "N/A")),
            ("入模特征数", feature_count or run_config.get("actual_feature_count", "N/A")),
            ("Best iteration", run_config.get("best_iteration", "N/A")),
            ("随机种子", run_config.get("random_seed", "N/A")),
        ],
    )

    row = _write_kv_section(
        ws,
        row,
        "Train / Valid 核心效果",
        [
            ("Train 样本数", metrics.get("train_samples", run_config.get("train_samples", "N/A"))),
            ("Valid 样本数", metrics.get("valid_samples", run_config.get("valid_samples", "N/A"))),
            ("Train bad rate", metrics.get("train_bad_rate", run_config.get("train_bad_rate", "N/A"))),
            ("Valid bad rate", metrics.get("valid_bad_rate", run_config.get("valid_bad_rate", "N/A"))),
            ("Train AUC", metrics.get("train_auc", "N/A")),
            ("Valid AUC", metrics.get("valid_auc", "N/A")),
            ("Train KS", metrics.get("train_ks", "N/A")),
            ("Valid KS", metrics.get("valid_ks", "N/A")),
            ("AUC Gap", metrics.get("auc_gap", "N/A")),
            ("训练耗时秒", metrics.get("train_time_seconds", "N/A")),
        ],
    )

    if eval_summary:
        row = _write_kv_section(
            ws,
            row,
            "评估范围",
            [
                ("总样本数", eval_summary.get("n_total_samples", "N/A")),
                ("评估分数", ", ".join(eval_summary.get("score_columns_evaluated", []))),
                ("评估切分", ", ".join(eval_summary.get("splits_evaluated", []))),
            ],
        )

    for title, path in [
        ("样本切分分布", _first_existing(sample_dir / "sample_split_summary.csv", input_dir / "sample_split_summary.csv")),
        ("标签分布", _first_existing(sample_dir / "label_distribution.csv", input_dir / "label_distribution.csv")),
        ("客群分布", _first_existing(sample_dir / "segment_distribution.csv", input_dir / "segment_distribution.csv")),
    ]:
        frame = _read_csv(path) if path else None
        if frame is not None and not frame.empty:
            row = _write_table(ws, row, title, frame)

    row = _write_kv_section(
        ws,
        row,
        "口径提示",
        [
            ("风险观察", "MOB1/MOB3 历史风险定义仍需确认，当前风险表不等同于历史正式口径"),
            ("缺失补充", "详见 reports/model_report_missing_results.md"),
        ],
    )


def _build_features_sheet(ws, *, train_dir: Path, feature_dir: Path) -> None:
    importance = _read_csv(train_dir / "feature_importance.csv")
    drop_detail = _read_csv(train_dir / "feature_drop_detail.csv")
    availability = _read_csv(feature_dir / "feature_availability.csv")
    final_features = _read_feature_list(train_dir, feature_dir)

    row = 1
    if importance is not None and not importance.empty:
        enriched = importance.copy()
        if drop_detail is not None and not drop_detail.empty:
            keep_cols = [col for col in ["feature", "non_null_rate", "unique_count", "drop_reason"] if col in drop_detail.columns]
            enriched = enriched.merge(drop_detail[keep_cols], on="feature", how="left")
        if availability is not None and not availability.empty:
            keep_cols = [col for col in ["feature", "in_feather", "dtype"] if col in availability.columns]
            enriched = enriched.merge(availability[keep_cols], on="feature", how="left")
        row = _write_table(ws, row, "Top 20 重要变量", enriched.head(20))
        row = _write_table(ws, row, "完整入模变量明细", enriched)

    if final_features:
        feature_frame = pd.DataFrame({"index": range(1, len(final_features) + 1), "feature": final_features})
        row = _write_table(ws, row, "最终特征清单", feature_frame)

    if drop_detail is not None and "drop_reason" in drop_detail.columns:
        dropped = drop_detail[drop_detail["drop_reason"].notna() & (drop_detail["drop_reason"].astype(str) != "")]
        if not dropped.empty:
            _write_table(ws, row, "被删除特征", dropped)


def _build_screening_params_sheet(ws, *, train_dir: Path, feature_dir: Path) -> None:
    stage_summary = _read_json(feature_dir / "feature_stage_summary.json")
    run_config = _read_json(train_dir / "run_config.json")
    params = run_config.get("params", {}) if isinstance(run_config.get("params"), dict) else {}

    row = 1
    if stage_summary:
        row = _write_table(ws, row, "变量筛选过程", _screening_steps_frame(stage_summary))
        row = _write_kv_section(
            ws,
            row,
            "筛选结果摘要",
            [
                ("筛选方法", stage_summary.get("filtering_method", "N/A")),
                ("数据可用特征数", stage_summary.get("features_available_in_data", "N/A")),
                ("缺失特征数", stage_summary.get("missing_features", "N/A")),
                ("潜在泄露提示", stage_summary.get("potential_leakage_flags", "无")),
            ],
        )

    if params:
        row = _write_table(ws, row, "LightGBM 参数", pd.DataFrame({"parameter": list(params.keys()), "value": list(params.values())}))

    _write_kv_section(
        ws,
        row,
        "训练配置",
        [
            ("experiment", run_config.get("experiment", "N/A")),
            ("data_source", run_config.get("data_source", "N/A")),
            ("candidate_feature_count", run_config.get("candidate_feature_count", "N/A")),
            ("actual_feature_count", run_config.get("actual_feature_count", "N/A")),
        ],
    )


def _build_monthly_effect_sheet(ws, *, eval_dir: Path) -> None:
    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    benchmark = _read_csv(eval_dir / "benchmark_uplift.csv")
    row = 1

    if monthly is not None and not monthly.empty:
        for final_flag in _ordered_values(monthly["final_flag"].dropna().unique().tolist()):
            subset = monthly[monthly["final_flag"] == final_flag].copy()
            if subset.empty:
                continue
            title = f"在全客群 {final_flag} 样本效果"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame(subset, row_col="mdl_month", metric="ks", row_label="30天发起Y"),
                right_frame=_metric_comparison_frame(subset, row_col="mdl_month", metric="auc", row_label="30天发起Y"),
            )

    if segment is not None and not segment.empty:
        row = _write_note(ws, row, "以下分客群表为切片效果，不代表已训练分客群专属模型；每张小表只横向对比一个指标。")
        for segment_name in ["老户次新", "老户", "次新", "流失户", "全客群"]:
            subset = segment[segment["segment"] == segment_name].copy()
            if subset.empty:
                continue
            title = f"在{segment_name}效果"
            row = _write_metric_table_pair(
                ws,
                row,
                title=title,
                left_frame=_metric_comparison_frame(subset, row_col="final_flag", metric="ks", row_label="样本"),
                right_frame=_metric_comparison_frame(subset, row_col="final_flag", metric="auc", row_label="样本"),
            )

    if benchmark is not None and not benchmark.empty:
        row = _write_table(ws, row, "历史版本提升摘要", benchmark)


def _build_sloping_sheet(ws, *, eval_dir: Path) -> None:
    row = 1
    for segment_label, segment_key in SEGMENT_FILES.items():
        row = _write_note(ws, row, f"2025-12 至 2026-01 30天发起 OOT/OOS：在{segment_label}效果")
        table_row = row
        max_end = row
        col = 1
        for score_column in SCORE_COLUMNS:
            frame = _read_csv(eval_dir / f"decile_lift_{segment_key}_{score_column}.csv")
            if frame is None and score_column == "model_score":
                frame = _read_csv(eval_dir / f"decile_lift_{segment_key}.csv")
            if frame is None or frame.empty:
                continue
            end = _write_table(
                ws,
                table_row,
                VERSION_LABELS.get(score_column, score_column),
                _sloping_display_frame(frame),
                start_col=col,
                apply_color_scale=False,
                plain=True,
            )
            max_end = max(max_end, end)
            col = 8 if col == 1 else 1
            if col == 1:
                table_row = max_end
        row = max(max_end + 1, table_row + 1)


def _build_intent_risk_sheet(ws, *, eval_dir: Path) -> None:
    row = 1
    row = _write_note(
        ws,
        row,
        "当前意愿资产交叉 artifact 缺少老户/流失户、score version、final_flag 和金额风险 x 资产评级维度；下方仅展示当前可用的全量观察口径，待补口径见 missing 文档。",
    )
    distribution = _read_csv(eval_dir / "intent_zc_distribution.csv")
    amount_risk = _read_csv(eval_dir / "intent_zc_amount_risk.csv")
    head_risk = _read_csv(eval_dir / "intent_zc_headcount_risk.csv")

    if distribution is not None and not distribution.empty:
        row = _write_table(ws, row, "当前可用全量观察：占比（意愿评级 x 资产评级）", _intent_sum_matrix(distribution, "pct"))
        row = _write_table(ws, row, "当前可用全量观察：30天发起率（意愿评级 x 资产评级）", _intent_rate_matrix(distribution, "bad", "n_samples"))
    if head_risk is not None and not head_risk.empty:
        row = _write_table(ws, row, "当前可用全量观察：人头风险率（意愿评级 x 资产评级）", _intent_rate_matrix(head_risk, "head_risk_count", "n_samples"))
    if amount_risk is not None and not amount_risk.empty:
        display = amount_risk.rename(
            columns={
                "intent_level": "意愿",
                "n_samples": "样本数",
                "total_principal": "本金金额",
                "total_overdue": "逾期金额",
                "amount_overdue_rate": "金额逾期率",
                "head_risk_count": "人头风险数",
                "head_risk_rate": "人头风险率",
            }
        )
        row = _write_table(ws, row, "当前可用全量观察：金额风险（仅意愿维度）", display)

    target = pd.DataFrame(
        [
            ("老户", "占比", "资产评级 x 意愿评级，含行/列 sum"),
            ("老户", "30天发起率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("老户", "新增订单3期金额逾期率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("流失户", "占比", "资产评级 x 意愿评级，含行/列 sum"),
            ("流失户", "30天发起率", "资产评级 x 意愿评级，含行/列加权整体"),
            ("流失户", "新增订单3期金额逾期率", "资产评级 x 意愿评级，含行/列加权整体"),
        ],
        columns=["待补客群", "待补指标", "目标矩阵"],
    )
    _write_table(ws, row, "待补矩阵口径", target, apply_color_scale=False)


def _build_stability_sheet(ws, *, eval_dir: Path) -> None:
    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    row = 1
    if psi is None or psi.empty:
        return

    if "score_column" in psi.columns:
        model_psi = psi[psi["score_column"] == "model_score"].copy()
        if model_psi.empty:
            model_psi = psi.copy()
    else:
        model_psi = psi.copy()

    display = model_psi.rename(columns={"month": "月份", "psi": "PSI", "n_samples": "样本数", "score_column": "分数"})
    row = _write_table(ws, row, "本轮模型月度 PSI", display)
    row = _write_note(ws, row, "当前 artifact 只有本轮模型月度 PSI 汇总，缺少每个分箱的跨月占比和发起率变化；待补口径见 missing 文档。")
    target = pd.DataFrame(
        [
            ("model_score", "月份 x 分箱", "score_bin/decile", "样本量、占比、30天发起率、PSI component"),
            ("model_score", "基准月 vs 观察月", "score_bin/decile", "占比差异、发起率差异"),
        ],
        columns=["分数", "期望粒度", "分箱字段", "期望指标"],
    )
    _write_table(ws, row, "待补稳定性分箱明细口径", target, apply_color_scale=False)


def _write_table(
    ws,
    start_row: int,
    title: str,
    frame: pd.DataFrame,
    *,
    source_note: str | None = None,
    start_col: int = 1,
    apply_color_scale: bool = True,
    plain: bool = False,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    frame = _clean_frame(frame)
    if frame.empty:
        return start_row

    ncols = max(len(frame.columns), 1)
    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=start_col + ncols - 1)
    title_cell = ws.cell(row=title_row, column=start_col)
    title_cell.value = title
    title_cell.font = Font(name="楷体", size=12, bold=True)
    if not plain:
        title_cell.fill = PatternFill(start_color="E7F4F2", end_color="E7F4F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = title_row + 1
    if source_note:
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col + ncols - 1)
        note_cell = ws.cell(row=header_row, column=start_col)
        note_cell.value = source_note
        note_cell.font = Font(name="楷体", size=10, italic=True, color="666666")
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_row += 1

    for col_offset, col_name in enumerate(frame.columns):
        cell = ws.cell(row=header_row, column=start_col + col_offset)
        cell.value = col_name
        cell.font = Font(name="楷体", size=10, bold=True)
        if not plain:
            cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row_data in enumerate(frame.itertuples(index=False), 1):
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=header_row + row_offset, column=start_col + col_offset)
            column_name = str(frame.columns[col_offset])
            cell.value = _excel_value(value, column_name)
            cell.number_format = _number_format(column_name, cell.value)
            cell.alignment = _alignment(column_name)

    end_row = header_row + len(frame)
    _style_region(ws, title_row, end_row, ncols, start_col=start_col)
    if apply_color_scale:
        _apply_table_color_scales(ws, header_row, end_row, frame, start_col=start_col)
    _update_column_widths(ws, frame, ncols, start_col=start_col)

    for col_idx in range(start_col, start_col + ncols):
        ws.column_dimensions[get_column_letter(col_idx)].bestFit = False

    return end_row + 3


def _write_metric_table_pair(ws, start_row: int, *, title: str, left_frame: pd.DataFrame, right_frame: pd.DataFrame) -> int:
    left_end = _write_comparison_table(ws, start_row, title, "KS", left_frame, start_col=1)
    right_end = _write_comparison_table(ws, start_row, title, "AUC", right_frame, start_col=9)
    return max(left_end, right_end)


def _write_comparison_table(
    ws,
    start_row: int,
    title: str,
    metric_label: str,
    frame: pd.DataFrame,
    *,
    start_col: int,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    frame = _clean_frame(frame)
    ncols = len(frame.columns)
    title_row = start_row
    metric_row = title_row + 1
    header_row = title_row + 2

    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=start_col + ncols - 1)
    title_cell = ws.cell(title_row, start_col)
    title_cell.value = title
    title_cell.font = Font(name="楷体", size=12, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=metric_row, start_column=start_col + 1, end_row=metric_row, end_column=start_col + ncols - 1)
    ws.cell(metric_row, start_col).value = frame.columns[0]
    ws.cell(metric_row, start_col).font = Font(name="楷体", size=10, bold=True)
    ws.cell(metric_row, start_col).alignment = Alignment(horizontal="center", vertical="center")
    metric_cell = ws.cell(metric_row, start_col + 1)
    metric_cell.value = metric_label
    metric_cell.font = Font(name="楷体", size=11, bold=True)
    metric_cell.alignment = Alignment(horizontal="center", vertical="center")
    metric_cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")

    for col_offset, col_name in enumerate(frame.columns):
        cell = ws.cell(header_row, start_col + col_offset)
        cell.value = col_name if col_offset > 0 else ""
        cell.font = Font(name="楷体", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="F3F6F6", end_color="F3F6F6", fill_type="solid")

    for row_offset, row_data in enumerate(frame.itertuples(index=False), 1):
        for col_offset, value in enumerate(row_data):
            column_name = str(frame.columns[col_offset])
            cell = ws.cell(header_row + row_offset, start_col + col_offset)
            cell.value = _excel_value(value, column_name)
            cell.number_format = _number_format(column_name, cell.value)
            cell.alignment = _alignment(column_name)

    end_row = header_row + len(frame)
    _style_region(ws, title_row, end_row, ncols, start_col=start_col)
    _apply_color_scale(ws, header_row + 1, end_row, start_col + 1, start_col + ncols - 1, prefer_high=True)
    _update_column_widths(ws, frame, ncols, start_col=start_col)
    return end_row + 3


def _write_kv_section(ws, start_row: int, title: str, pairs: list[tuple[str, Any]]) -> int:
    frame = pd.DataFrame(pairs, columns=["项目", "内容"])
    return _write_table(ws, start_row, title, frame)


def _write_note(ws, start_row: int, text: str) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    cell = ws.cell(row=start_row, column=1)
    cell.value = text
    cell.font = Font(name="楷体", size=10, italic=True, color="9C6500")
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _style_region(ws, start_row, start_row, 8)
    return start_row + 2


def _screening_steps_frame(stage_summary: dict[str, Any]) -> pd.DataFrame:
    process = _read_feature_screening_process()
    if process:
        rows = process.get("screening_rows", [])
        if rows:
            return pd.DataFrame(
                {
                    "步骤": [row.get("step") for row in rows],
                    "筛选方法": [row.get("method") for row in rows],
                    "剩余变量个数": [row.get("remaining_features") for row in rows],
                    "来源": [row.get("source") for row in rows],
                }
            )
    rows = [
        ("原始候选变量总数", "original_candidate_features"),
        ("分表基础预筛：缺失率、相关性、IV", "d01_kept_features"),
        ("稳定性筛选：DEV vs OOT PSI", "d02_kept_features"),
        ("Feather观察样本可用特征", "feather_available_features"),
        ("全局相关性去重：按单变量AUC保留更强特征", "after_global_corr"),
        ("随机噪声重要性筛选：剔除弱于噪声的真实特征", "after_d03_random_importance"),
        ("空标签重要性筛选：保留显著高于空标签分布的特征", "after_d04_null_importance"),
        ("最终训练特征", "final_training_features"),
    ]
    return pd.DataFrame(
        {
            "步骤": range(1, len(rows) + 1),
            "筛选方法": [label for label, _ in rows],
            "剩余变量个数": [stage_summary.get(key, "N/A") for _, key in rows],
        }
    )


def _metric_comparison_frame(frame: pd.DataFrame, *, row_col: str, metric: str, row_label: str) -> pd.DataFrame:
    cols = [row_col]
    rename = {row_col: row_label}
    for score_column in SCORE_COLUMNS:
        source_col = f"{score_column}_{metric}"
        if source_col in frame.columns:
            cols.append(source_col)
            rename[source_col] = VERSION_LABELS.get(score_column, score_column)
    display = frame[cols].copy().rename(columns=rename)
    return display


def _sloping_display_frame(frame: pd.DataFrame) -> pd.DataFrame:
    display = frame.copy().sort_values("decile", ascending=True).reset_index(drop=True)
    required = {"decile", "n_samples", "bad"}
    if not required.issubset(display.columns):
        return pd.DataFrame(
            {
                "分组": [f"{int(row.decile):03d}" for row in display.itertuples(index=False)],
                "占比": display.get("pct"),
                "累计发起率": display.get("cum_bad_rate"),
                "累计lift": display.get("cum_lift"),
                "剩余发起率": display.get("remaining_bad_rate"),
                "剩余lift": display.get("remaining_lift"),
            }
        )

    n_samples = pd.to_numeric(display["n_samples"], errors="coerce").fillna(0)
    bad = pd.to_numeric(display["bad"], errors="coerce").fillna(0)
    total_n = float(n_samples.sum())
    total_bad = float(bad.sum())
    total_rate = total_bad / total_n if total_n else 0.0

    cum_n = n_samples.cumsum()
    cum_bad = bad.cumsum()
    remaining_n = total_n - cum_n
    remaining_bad = total_bad - cum_bad
    cum_rate = cum_bad.divide(cum_n).where(cum_n > 0, 0)
    remaining_rate = remaining_bad.divide(remaining_n).where(remaining_n > 0, 0)

    return pd.DataFrame(
        {
            "分组": [f"{int(row.decile):03d}" for row in display.itertuples(index=False)],
            "占比": n_samples / total_n if total_n else 0,
            "累计发起率": cum_rate,
            "累计lift": cum_rate / total_rate if total_rate else 0,
            "剩余发起率": remaining_rate,
            "剩余lift": remaining_rate / total_rate if total_rate else 0,
        }
    )


def _intent_sum_matrix(frame: pd.DataFrame, value_col: str) -> pd.DataFrame:
    pivot = frame.pivot_table(index="intent_level", columns="zc_level", values=value_col, aggfunc="first").reset_index()
    pivot.columns.name = None
    pivot = pivot.rename(columns={"intent_level": "意愿"})
    pivot = _sort_intent_matrix(pivot)
    value_cols = [col for col in pivot.columns if col != "意愿"]
    if value_cols:
        pivot["sum"] = pivot[value_cols].sum(axis=1)
        total = {"意愿": "sum"}
        total.update({col: pivot[col].sum() for col in value_cols})
        total["sum"] = pivot["sum"].sum()
        pivot = pd.concat([pivot, pd.DataFrame([total])], ignore_index=True)
    return pivot


def _intent_rate_matrix(frame: pd.DataFrame, numerator_col: str, denominator_col: str) -> pd.DataFrame:
    rows = []
    work = frame.copy()
    work[numerator_col] = pd.to_numeric(work[numerator_col], errors="coerce").fillna(0)
    work[denominator_col] = pd.to_numeric(work[denominator_col], errors="coerce").fillna(0)
    zc_values = sorted(work["zc_level"].dropna().unique().tolist(), key=lambda value: str(value))
    for intent in ["低意愿", "中意愿", "高意愿"]:
        sub = work[work["intent_level"] == intent]
        if sub.empty:
            continue
        row: dict[str, Any] = {"意愿": intent}
        for zc_value in zc_values:
            cell = sub[sub["zc_level"] == zc_value]
            numerator = float(cell[numerator_col].sum())
            denominator = float(cell[denominator_col].sum())
            row[zc_value] = numerator / denominator if denominator else None
        numerator = float(sub[numerator_col].sum())
        denominator = float(sub[denominator_col].sum())
        row["sum"] = numerator / denominator if denominator else None
        rows.append(row)
    if not rows:
        return pd.DataFrame()

    total: dict[str, Any] = {"意愿": "sum"}
    for zc_value in zc_values:
        cell = work[work["zc_level"] == zc_value]
        numerator = float(cell[numerator_col].sum())
        denominator = float(cell[denominator_col].sum())
        total[zc_value] = numerator / denominator if denominator else None
    total["sum"] = float(work[numerator_col].sum()) / float(work[denominator_col].sum()) if float(work[denominator_col].sum()) else None
    rows.append(total)
    return pd.DataFrame(rows)


def _sort_intent_matrix(frame: pd.DataFrame) -> pd.DataFrame:
    if "意愿" not in frame.columns:
        return frame
    order = {"低意愿": 0, "中意愿": 1, "高意愿": 2, "sum": 3}
    sorted_frame = frame.copy()
    sorted_frame["_sort"] = sorted_frame["意愿"].map(order).fillna(99)
    sorted_frame = sorted_frame.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    return sorted_frame


def _ordered_values(values: list[Any]) -> list[Any]:
    preferred = ["DEV", "DEV-OOS", "OOT", "OOT-OOS"]
    ordered = [value for value in preferred if value in values]
    ordered.extend([value for value in values if value not in ordered])
    return ordered


def _read_feature_screening_process() -> dict[str, Any]:
    repo_root = Path(__file__).resolve().parents[3]
    for path in [
        repo_root / "projects" / "2026-05-fujie-gcard-v1" / "reports" / "feature_screening_process.json",
        repo_root
        / "projects"
        / "2026-05-fujie-gcard-v1"
        / "runs"
        / "2026-05-imported-feature-screening"
        / "feature_selection"
        / "feature_screening_process.json",
    ]:
        if path.exists():
            return _read_json(path)
    return {}


def _clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    cleaned = frame.copy()
    cleaned = cleaned.where(pd.notna(cleaned), None)
    return cleaned


def _excel_value(value: Any, column_name: str) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, str):
        return value
    if _is_count_column(column_name):
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if isinstance(value, float):
        return float(value)
    return value


def _number_format(column_name: str, value: Any) -> str:
    if value is None:
        return "General"
    if _is_count_column(column_name):
        return "#,##0"
    lowered = column_name.lower()
    if any(token in lowered for token in ["amount", "principal", "overdue", "amt"]):
        return "#,##0.000"
    if _is_percent_column(column_name):
        return "0.0%"
    if isinstance(value, float):
        return "0.000"
    return "General"


def _alignment(column_name: str):
    from openpyxl.styles import Alignment

    if _is_count_column(column_name) or isinstance(column_name, str) and _looks_metric_column(column_name):
        return Alignment(horizontal="right", vertical="center", wrap_text=True)
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _is_count_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    return any(token in lowered for token in ["count", "samples", "positive", "bad", "split", "unique"]) and not any(
        token in lowered for token in ["rate", "ratio", "pct", "lift", "auc", "ks"]
    )


def _is_percent_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    name = str(column_name)
    if any(token in lowered for token in ["auc", "ks", "lift"]):
        return False
    return any(token in lowered for token in ["rate", "ratio", "pct"]) or any(token in name for token in ["率", "占比"])


def _looks_metric_column(column_name: str) -> bool:
    lowered = str(column_name).lower()
    return any(
        token in lowered
        for token in [
            "auc",
            "ks",
            "rate",
            "ratio",
            "pct",
            "lift",
            "psi",
            "gain",
            "score",
            "amount",
            "principal",
            "overdue",
            "value",
            "发起率",
            "占比",
            "金额逾期率",
            "人头风险率",
        ]
    )


def _style_region(ws, start_row: int, end_row: int, ncols: int, *, start_col: int = 1) -> None:
    from openpyxl.styles import Border, Font, Side

    thin = Side(style="thin", color="D9E0E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=start_col + ncols - 1):
        for cell in row:
            cell.border = border
            if cell.font is None or cell.font.name is None:
                cell.font = Font(name="楷体", size=10)


def _apply_table_color_scales(ws, header_row: int, end_row: int, frame: pd.DataFrame, *, start_col: int = 1) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    if end_row <= header_row:
        return

    for col_offset, column_name in enumerate(frame.columns):
        lowered = str(column_name).lower()
        if not _should_color_scale(lowered):
            continue
        col_letter = get_column_letter(start_col + col_offset)
        cell_range = f"{col_letter}{header_row + 1}:{col_letter}{end_row}"
        if any(token in lowered for token in ["uplift", "提升"]):
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFFFF",
                end_type="max",
                end_color="63BE7B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            )
        ws.conditional_formatting.add(cell_range, rule)


def _apply_color_scale(ws, start_row: int, end_row: int, start_col: int, end_col: int, *, prefer_high: bool) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    if end_row < start_row or end_col < start_col:
        return
    if prefer_high:
        start_color, end_color = "63BE7B", "F8696B"
    else:
        start_color, end_color = "F8696B", "63BE7B"
    rule = ColorScaleRule(
        start_type="min",
        start_color=start_color,
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color=end_color,
    )
    ws.conditional_formatting.add(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}", rule)


def _should_color_scale(lowered_column_name: str) -> bool:
    return any(
        token in lowered_column_name
        for token in [
            "uplift",
            "cum_lift",
            "remaining_lift",
            "bad_rate",
            "head_risk_rate",
            "amount_overdue_rate",
            "psi",
            "占比",
            "发起率",
            "金额逾期率",
            "人头风险率",
        ]
    )


def _update_column_widths(ws, frame: pd.DataFrame, ncols: int, *, start_col: int = 1) -> None:
    from openpyxl.utils import get_column_letter

    for col_offset in range(ncols):
        col_name = str(frame.columns[col_offset])
        max_len = len(col_name)
        sample_values = frame.iloc[:80, col_offset].tolist()
        for value in sample_values:
            if value is not None and not pd.isna(value):
                max_len = max(max_len, len(str(value)))
        if any(token in col_name.lower() for token in ["feature", "desc", "data_source", "内容"]):
            width = min(max(max_len * 1.1, 18), 52)
        elif _looks_metric_column(col_name) or _is_count_column(col_name):
            width = min(max(max_len * 0.85, 10), 16)
        else:
            width = min(max(max_len * 1.0, 10), 24)
        ws.column_dimensions[get_column_letter(start_col + col_offset)].width = width


def _finalize_sheet(ws) -> None:
    from copy import copy

    from openpyxl.styles import Alignment, Font

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_format.defaultRowHeight = 18
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                font = copy(cell.font)
                if font.name is None:
                    font.name = "楷体"
                    cell.font = font
                alignment = copy(cell.alignment) if cell.alignment else Alignment()
                if alignment.vertical is None:
                    alignment.vertical = "center"
                if alignment.wrap_text is None:
                    alignment.wrap_text = True
                cell.alignment = alignment
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = "E7F4F2"
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws["A1"] = "暂无可用数据"
        ws["A1"].font = Font(name="楷体", size=11, italic=True)


def _write_model_reports(
    *,
    output_path: Path,
    train_dir: Path,
    eval_dir: Path,
    feature_dir: Path,
    sample_dir: Path,
) -> tuple[Path, Path]:
    run_config = _read_json(train_dir / "run_config.json")
    metrics = _read_json(train_dir / "metrics_train_valid.json")
    stage_summary = _read_json(feature_dir / "feature_stage_summary.json")
    screening_process = _read_feature_screening_process()
    overall = _read_csv(eval_dir / "overall_metrics.csv")
    benchmark = _read_csv(eval_dir / "benchmark_uplift.csv")
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    importance = _read_csv(train_dir / "feature_importance.csv")
    sample_split = _read_csv(sample_dir / "sample_split_summary.csv")

    md_path = output_path.with_name("model_report.md")
    html_path = output_path.with_name("model_report.html")

    final_features = stage_summary.get("final_training_features", run_config.get("actual_feature_count", "N/A"))
    valid_auc = _fmt_metric(metrics.get("valid_auc"))
    valid_ks = _fmt_metric(metrics.get("valid_ks"))
    auc_gap = _fmt_metric(metrics.get("auc_gap"))
    oot_oos = _row_by_value(benchmark, "final_flag", "OOT-OOS")

    lines = [
        "# 复借G卡模型报告",
        "",
        f"生成日期：{time.strftime('%Y-%m-%d')}",
        "",
        "## 一、模型描述",
        "",
        f"- 模型目标：预测观察日后 30 天内是否发起，标签字段为 `{run_config.get('label_column', 'ftr_30d_ord_flag')}`。",
        f"- 建模样本：训练集 {_fmt_list(run_config.get('train_values', []))}，验证集 {_fmt_list(run_config.get('valid_values', []))}，OOS {_fmt_list(run_config.get('oos_values', []))}。",
        f"- 算法：{run_config.get('algorithm', 'N/A')}；最终入模变量 {final_features} 个；best iteration {run_config.get('best_iteration', 'N/A')}。",
        f"- 验证集效果：AUC {valid_auc}，KS {valid_ks}，Train/Valid AUC gap {auc_gap}。",
        "",
        "## 二、变量筛选过程",
        "",
    ]
    if screening_process.get("feature_select_v2_alignment", {}).get("summary"):
        lines.append(f"- {screening_process['feature_select_v2_alignment']['summary']}")
        lines.append("")
    screening_rows = screening_process.get("screening_rows") or _screening_steps_frame(stage_summary).to_dict("records")
    lines.extend(
        _markdown_table(
            pd.DataFrame(screening_rows).rename(
                columns={"step": "步骤", "method": "筛选方法", "remaining_features": "剩余变量个数", "source": "来源"}
            )
        )
    )
    lines.extend(
        [
            "",
            "## 三、核心效果与历史版本对比",
            "",
            _metric_sentence("OOT-OOS", oot_oos),
            "",
        ]
    )
    if overall is not None:
        lines.extend(_markdown_table(overall[["final_flag", "n_samples", "positive", "bad_rate", "model_score_auc", "model_score_ks"]]))
        lines.append("")
    if benchmark is not None:
        display_cols = [
            "final_flag",
            "model_score_auc",
            "model_score_ks",
            "ks_uplift_vs_gcard_v2",
            "ks_uplift_vs_gcard_v4",
            "ks_uplift_vs_gcard_v5",
            "ks_uplift_vs_gcard_v6",
        ]
        lines.extend(_markdown_table(benchmark[[col for col in display_cols if col in benchmark.columns]]))
        lines.append("")

    lines.extend(["## 四、分客群效果", ""])
    if segment is not None:
        seg_cols = ["segment", "final_flag", "n_samples", "bad_rate", "model_score_auc", "model_score_ks", "ks_uplift_vs_gcard_v2"]
        lines.extend(_markdown_table(segment[[col for col in seg_cols if col in segment.columns]].head(20)))
        lines.append("")

    lines.extend(
        [
            "## 五、模型 sloping、意愿交叉风险与稳定性",
            "",
            "- sloping 详见 Excel 中 `模型效果-模型sloping`，已按全客群、老户次新、流失户分别横向对比本轮模型和历史 G 卡版本；累计和剩余 lift 按参考文档口径从低分尾部开始累计。",
            "- 意愿交叉风险详见 Excel 中 `模型效果-意愿交叉风险（DEV-OOS）`。当前 artifact 缺少老户/流失户、score version、final_flag 和金额风险 x 资产评级维度，不伪造缺失矩阵。",
        ]
    )
    if psi is not None and not psi.empty:
        if "score_column" in psi.columns:
            psi = psi[psi["score_column"] == "model_score"].copy()
        max_psi = psi.sort_values("psi", ascending=False).head(5)
        lines.append("- 本轮模型 PSI 最高的 5 个观测如下；分箱占比变化明细需补充产出后回填：")
        lines.extend(_markdown_table(max_psi))
    lines.extend(
        [
            "",
            "## 六、重要变量",
            "",
        ]
    )
    if importance is not None:
        lines.extend(_markdown_table(importance.head(15)))
        lines.append("")
    lines.extend(
        [
            "## 七、待补充事项",
            "",
            "- 历史文档中的变量分布/分箱图、变量中文描述与业务标签、MOB1/MOB3 历史风险精确定义仍需在另一环境补充计算。",
            "- 详见 `model_report_missing_results.md`。",
        ]
    )

    markdown = "\n".join(lines).rstrip() + "\n"
    md_path.write_text(markdown, encoding="utf-8")
    html_path.write_text(_markdown_to_simple_html(markdown), encoding="utf-8")
    return md_path, html_path


def _fmt_list(values: Any) -> str:
    if isinstance(values, list):
        return "、".join(str(value) for value in values) if values else "N/A"
    return str(values)


def _metric_sentence(label: str, row: dict[str, Any]) -> str:
    if not row:
        return f"- {label} 暂无可用 benchmark 指标。"
    return (
        f"- {label}：本轮模型 AUC {_fmt_metric(row.get('model_score_auc'))}、KS {_fmt_metric(row.get('model_score_ks'))}；"
        f"相对 G卡V2 KS 提升 {_fmt_metric(row.get('ks_uplift_vs_gcard_v2'))}，"
        f"相对 G卡V4 KS 提升 {_fmt_metric(row.get('ks_uplift_vs_gcard_v4'))}，"
        f"相对 G卡V6 KS 提升 {_fmt_metric(row.get('ks_uplift_vs_gcard_v6'))}。"
    )


def _row_by_value(frame: pd.DataFrame | None, column: str, value: Any) -> dict[str, Any]:
    if frame is None or frame.empty or column not in frame.columns:
        return {}
    matched = frame[frame[column] == value]
    if matched.empty:
        return {}
    return matched.iloc[0].to_dict()


def _markdown_table(frame: pd.DataFrame, *, limit: int = 20) -> list[str]:
    if frame.empty:
        return ["暂无可用数据"]
    display = frame.head(limit).copy()
    display = display.rename(columns={col: str(col) for col in display.columns})
    rows = []
    headers = [str(col) for col in display.columns]
    rows.append("| " + " | ".join(headers) + " |")
    rows.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for item in display.itertuples(index=False):
        rows.append("| " + " | ".join(_fmt_markdown_cell(value) for value in item) + " |")
    if len(frame) > limit:
        rows.append(f"\n> 仅展示前 {limit} 行，完整明细见 Excel。")
    return rows


def _fmt_markdown_cell(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.3f}"
    return str(value).replace("|", "\\|")


def _fmt_metric(value: Any) -> str:
    if value is None or value == "N/A":
        return "N/A"
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)


def _markdown_to_simple_html(markdown: str) -> str:
    lines = markdown.splitlines()
    html_lines = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head>",
        '<meta charset="utf-8">',
        "<title>复借G卡模型报告</title>",
        "<style>",
        "body{font-family:Arial,'Microsoft YaHei',sans-serif;margin:32px;color:#1f2933;line-height:1.55}",
        "h1{font-size:26px;margin-bottom:16px} h2{font-size:19px;margin-top:28px;border-bottom:1px solid #d9e0e3;padding-bottom:6px}",
        "table{border-collapse:collapse;margin:12px 0 20px 0;font-size:13px} th,td{border:1px solid #d9e0e3;padding:6px 8px;text-align:right} th{background:#f3f6f6} td:first-child,th:first-child{text-align:left}",
        "li{margin:4px 0} blockquote{color:#667085;border-left:3px solid #d9e0e3;padding-left:10px}",
        "</style>",
        "</head><body>",
    ]
    in_ul = False
    in_table = False
    for line in lines:
        if line.startswith("# "):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            html_lines.append(f"<h1>{escape(line[2:])}</h1>")
        elif line.startswith("## "):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            html_lines.append(f"<h2>{escape(line[3:])}</h2>")
        elif line.startswith("- "):
            if not in_ul:
                html_lines.append("<ul>")
                in_ul = True
            html_lines.append(f"<li>{escape(line[2:])}</li>")
        elif line.startswith("| ") and line.endswith(" |"):
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if all(cell == "---" for cell in cells):
                continue
            if not in_table:
                html_lines.append("<table>")
                in_table = True
                tag = "th"
            else:
                tag = "td"
            html_lines.append("<tr>" + "".join(f"<{tag}>{escape(cell)}</{tag}>" for cell in cells) + "</tr>")
        elif line.startswith("> "):
            if in_table:
                html_lines.append("</table>")
                in_table = False
            html_lines.append(f"<blockquote>{escape(line[2:])}</blockquote>")
        else:
            if in_table:
                html_lines.append("</table>")
                in_table = False
            if in_ul:
                html_lines.append("</ul>")
                in_ul = False
            if line.strip():
                html_lines.append(f"<p>{escape(line)}</p>")
    if in_table:
        html_lines.append("</table>")
    if in_ul:
        html_lines.append("</ul>")
    html_lines.append("</body></html>")
    return "\n".join(html_lines)


def _write_missing_results_doc(output_path: Path) -> Path:
    missing_path = output_path.with_name("model_report_missing_results.md")
    text = """# 复借G卡模型报告缺失结果清单

本文件只记录当前已注册 run artifact 无法可靠还原的内容，不伪造指标。

| 缺少字段/结果 | 期望粒度 | 建议产出文件 | 可填入目标 sheet |
|---|---|---|---|
| 历史文档中的图片化变量分布/分箱图 | 每个重要变量、每个分箱 | `reports/variable_bin_plots/*.png` 或 `evaluation/variable_bins.csv` | `重要变量` |
| 变量中文描述、业务标签 | feature 级别 | `feature_selection/feature_metadata.csv`，字段建议包含 `feature,desc,label` | `重要变量` |
| MOB1/MOB3 历史风险精确定义 | final_flag、segment、score、decile、intent_level、zc_level | `evaluation/mob_risk_metrics.csv` | `模型效果-模型sloping`、`模型效果-意愿交叉风险（DEV-OOS）` |
| sloping 分箱上下界 | `segment`、`score/version`、`decile`、`lower_bound`、`upper_bound`，需要能展示为 `001:(-inf, x]` 这类区间 | `evaluation/decile_lift_bins.csv`，或在各 `decile_lift_*.csv` 增加 `score_min,score_max` | `模型效果-模型sloping` |
| 明确 DEV-OOS 过滤后的意愿资产交叉结果 | `segment in (老户, 流失户)`、`score/version`、`intent_level`、`zc_level`，限定 `final_flag=DEV-OOS` | `evaluation/intent_zc_dev_oos_*.csv` | `模型效果-意愿交叉风险（DEV-OOS）` |
| 老户/流失户意愿资产占比矩阵 | `segment in (老户, 流失户)`、`score/version`、`intent_level`、`zc_level`，意愿按对应模型分等频三份；指标包含 `n_samples,sample_pct,row_pct,col_pct` 和行/列 sum | `evaluation/intent_zc_segment_distribution.csv` | `模型效果-意愿交叉风险（DEV-OOS）` |
| 老户/流失户意愿资产 30 天发起率矩阵 | `segment in (老户, 流失户)`、`score/version`、`intent_level`、`zc_level`；指标包含 `n_samples,ftr_30d_count,ftr_30d_rate` 和行/列加权整体 | `evaluation/intent_zc_segment_ftr_rate.csv` | `模型效果-意愿交叉风险（DEV-OOS）` |
| 老户/流失户意愿资产新增订单 3 期金额逾期率矩阵 | `segment in (老户, 流失户)`、`score/version`、`intent_level`、`zc_level`；指标包含 `total_principal,total_overdue,amount_overdue_rate` 和行/列加权整体 | `evaluation/intent_zc_segment_amount_risk.csv` | `模型效果-意愿交叉风险（DEV-OOS）` |
| 老户次新、流失户 OOT-OOS 客群 by 月模型效果 | `mdl_month`、`segment in (老户次新, 流失户)`、`final_flag=OOT-OOS`、`score/version`，指标包含 `n_samples, positive, bad_rate, AUC, KS` | `evaluation/monthly_segment_metrics.csv` 或 `evaluation/monthly_segment_metrics_oot_oos.csv` | `模型效果-每月效果` |
| 分客群训练模型与全客群模型的同口径对比 | segment、final_flag、score、AUC、KS、lift | `evaluation/segment_model_comparison.csv` | `模型效果-每月效果`、`模型效果-模型sloping` |
| 本轮模型分箱稳定性明细 | `score_column=model_score`、`month`、`score_bin/decile`；指标包含 `n_samples,pct,bad_rate,psi_component`，需能展示每个分箱占比随月份变化 | `evaluation/model_score_bin_distribution_by_month.csv` | `模型稳定性` |

补齐这些文件后，应继续通过 `jm report` 统一生成报告，避免人工改写 xlsx 中的计算结果。
"""
    missing_path.write_text(text, encoding="utf-8")
    return missing_path


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv(path: Path) -> pd.DataFrame | None:
    if not path.exists():
        return None
    return pd.read_csv(path, encoding="utf-8-sig")


def _first_existing(*paths: Path) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def _feature_count(train_dir: Path, feature_dir: Path) -> int:
    return len(_read_feature_list(train_dir, feature_dir))


def _read_feature_list(train_dir: Path, feature_dir: Path) -> list[str]:
    for path in [
        feature_dir / "final_features.txt",
        feature_dir / "feature_list.txt",
        train_dir / "actual_feature_list.txt",
    ]:
        if path.exists():
            return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    return []
