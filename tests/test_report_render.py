from pathlib import Path
import base64
import shutil

from openpyxl import load_workbook

from risk_model_workbench.cli import main
from risk_model_workbench.reporting.excel_report import REPORT_SHEETS, generate_excel_report


def test_report_scaffold():
    project = "projects/2026-05-fujie-gcard-v1"
    run_id = "pytest_report_scaffold"
    run_dir = Path(project) / "runs" / run_id
    try:
        main(["run", "init", "--project", project, "--workflow", "full_modeling", "--run-id", run_id, "--force"])
        assert main(["report", "--project", project, "--run-id", run_id]) == 0
        assert (run_dir / "reports" / "model_report.md").exists()
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)


def test_imported_excel_report_layout(tmp_path):
    run_dir = Path("projects/2026-05-fujie-gcard-v1/runs/2026-06-imported-gcard-main-lgbm")
    output_path = tmp_path / "model_report.xlsx"

    generate_excel_report(
        eval_dir=run_dir / "evaluation",
        train_dir=run_dir / "modeling" / "main_lgbm",
        input_dir=run_dir / "modeling_input",
        feature_dir=run_dir / "feature_selection",
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == REPORT_SHEETS
    assert output_path.with_name("model_report_missing_results.md").exists()
    assert output_path.with_name("model_report.md").exists()
    assert output_path.with_name("model_report.html").exists()

    description = workbook["模型描述"]
    assert _find_cell(description, "模型结论摘要") is None
    assert _find_cell(description, "样本切分分布") is not None
    assert _find_cell(description, "samples") is not None
    sample_cell = _cell_below_header(description, "samples")
    bad_rate_cell = _cell_below_header(description, "bad_rate")
    assert sample_cell.value == 3600000
    assert sample_cell.number_format == "#,##0"
    assert round(bad_rate_cell.value, 6) == round(550617 / 3600000, 6)
    assert bad_rate_cell.number_format == "0.0%"

    ws = workbook["模型效果-每月效果"]
    assert _find_cell(ws, "1、每月效果（OOS）") is not None
    assert _find_cell_contains(ws, "在 OOT-OOS 样本上老户次新客群上") is not None
    assert _find_cell_contains(ws, "对比G卡V5") is not None
    assert _find_cell_contains(ws, "对比G卡V6") is not None
    assert _find_cell_contains(ws, "未注册老户次新/流失户专属模型得分") is not None
    assert _find_cell(ws, "KS") is not None
    assert _find_cell(ws, "AUC") is not None
    assert _find_cell(ws, "本轮模型") is not None
    assert _find_cell(ws, "在全客群 OOS by月效果（DEV-OOS + OOT-OOS）") is not None
    assert _find_cell(ws, "在老户次新 OOS by月效果（DEV-OOS + OOT-OOS）") is not None
    assert _find_cell(ws, "在流失户 OOS by月效果（DEV-OOS + OOT-OOS）") is not None
    assert _find_cell(ws, "DEV-OOS 2025-06") is not None
    assert _find_cell(ws, "OOT-OOS 2026-01") is not None
    metric_cell = _cell_below_header(ws, "本轮模型")
    assert metric_cell is not None
    assert metric_cell.number_format == "0.000"
    assert not isinstance(metric_cell.value, str)

    screening = workbook["变量筛选过程和模型参数"]
    assert _find_cell(screening, "筛选方法") is not None
    assert _find_cell(screening, "特征初筛-质量规则：缺失率 < 0.95，相关性 < 0.80，IV >= 0.005") is not None

    sloping = workbook["模型效果-模型sloping"]
    assert _find_cell(sloping, "2、模型sloping") is not None
    assert _find_cell_contains(sloping, "高分10%分层") is not None
    for header in ["分组", "占比", "累计发起率", "累计lift", "剩余发起率", "剩余lift"]:
        assert _find_cell(sloping, header) is not None
    assert _find_cell_contains(sloping, "001:(-inf") is not None

    intent = workbook["模型效果-意愿交叉风险（DEV-OOS）"]
    assert _find_cell(intent, "3、意愿交叉风险（DEV-OOS）") is not None
    assert _find_cell_contains(intent, "高、中、低意愿评级") is not None
    assert _find_cell(intent, "老户 - 占比 - 本轮模型") is not None
    assert _find_cell(intent, "老户 - 占比 - G卡V6") is not None
    assert _find_cell(intent, "流失户 - 新增订单3期金额逾期率 - G卡V6") is not None

    stability = workbook["模型稳定性"]
    assert _find_cell(stability, "本轮模型分箱占比变化") is not None
    assert _find_cell_contains(stability, "001:") is not None

    woe = workbook["Top变量WOE"]
    assert _find_cell_contains(woe, "WOE charts require row-level feature values") is not None

    missing_text = output_path.with_name("model_report_missing_results.md").read_text(encoding="utf-8")
    assert "已补齐 — 历史版本横向对比" in missing_text
    assert "monthly_segment_metrics_oos_by_version.csv" in missing_text
    assert "不可补齐（3 项）" in missing_text
    report_text = output_path.with_name("model_report.md").read_text(encoding="utf-8")
    assert "1、每月效果（OOS）" in report_text
    assert "在老户次新 OOS by月效果（KS）" in report_text
    assert "2、模型sloping" in report_text
    assert "OOT-OOS 30天发起：在老户次新效果" in report_text
    assert "3、意愿交叉风险（DEV-OOS）" in report_text
    assert "老户 - 占比 - 本轮模型" in report_text

    assert len(workbook["模型效果-模型sloping"].conditional_formatting) == 0
    assert len(workbook["模型效果-意愿交叉风险（DEV-OOS）"].conditional_formatting) > 0
    assert len(workbook["模型稳定性"].conditional_formatting) > 0


def test_train_300_report_uses_current_run_training_features(tmp_path):
    run_dir = Path("projects/2026-05-fujie-gcard-v1/runs/20260615_train_300")
    output_path = tmp_path / "model_report.xlsx"

    generate_excel_report(
        eval_dir=run_dir / "evaluation",
        train_dir=run_dir / "modeling" / "main_lgbm",
        input_dir=run_dir / "modeling_input",
        feature_dir=run_dir / "feature_selection",
        output_path=output_path,
    )

    report_text = output_path.with_name("model_report.md").read_text(encoding="utf-8")
    report_html = output_path.with_name("model_report.html").read_text(encoding="utf-8")

    assert "最终入模变量 300 个" in report_text
    assert "当前 run 未登记独立特征初筛或特征精筛过程产物" in report_text
    assert "| 训练输入 |" in report_text
    assert "| 训练预处理 |" in report_text
    assert "| 最终入模 | LightGBM 实际入模变量数 | 300 |" in report_text
    assert "随机噪声重要性筛选：3轮" not in report_text
    assert "runs/feature_refine_feather/stage_summary.json" not in report_text

    assert "全客群 by月效果（KS）" in report_text
    assert "分客群整体效果（AUC）" in report_text
    assert "2、模型sloping" in report_text
    assert "3、意愿交叉风险（DEV-OOS）" in report_text
    assert "最终入模" in report_html
    assert "全客群 by月效果" in report_html

    workbook = load_workbook(output_path)
    screening = workbook["变量筛选过程和模型参数"]
    assert _find_cell(screening, "训练特征准备") is not None
    assert _find_cell(screening, "最终入模") is not None
    assert _cell_below_header(screening, "变量个数").value == 301


def test_report_embeds_top_feature_woe_sheet(tmp_path):
    eval_dir = tmp_path / "evaluation"
    train_dir = tmp_path / "modeling" / "main_lgbm"
    input_dir = tmp_path / "modeling_input"
    feature_dir = tmp_path / "feature_selection"
    for directory in [eval_dir, train_dir, input_dir, feature_dir]:
        directory.mkdir(parents=True)

    woe_dir = train_dir / "woe_top_features"
    image_dir = woe_dir / "images"
    image_dir.mkdir(parents=True)
    (woe_dir / "woe_top20_summary.csv").write_text(
        "\n".join(
            [
                "feature,rank,gain,split_importance,bin_order,bin_label,lower_bound,upper_bound,is_missing_bin,split_value,good,bad,total,bad_rate,pop_pct,woe,iv_component,status,skip_reason",
                "feature_a,1,30,5,0,Missing,,,True,DEV,10,1,11,0.0909,0.1,-1.2,0.02,ok,",
                "feature_a,1,30,5,1,\"(-inf, 1]\",-inf,1,False,DEV,5,5,10,0.5,0.2,0.5,0.01,ok,",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    _write_tiny_png(image_dir / "001_feature_a_WOE.png")

    output_path = tmp_path / "reports" / "model_report.xlsx"
    generate_excel_report(
        eval_dir=eval_dir,
        train_dir=train_dir,
        input_dir=input_dir,
        feature_dir=feature_dir,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    woe = workbook["Top变量WOE"]
    assert _find_cell(woe, "Top 1: feature_a") is not None
    assert _find_cell(woe, "Gain") is not None
    assert len(woe._images) == 1

    report_text = output_path.with_name("model_report.md").read_text(encoding="utf-8")
    assert "## 七、Top变量WOE" in report_text
    assert "feature_a" in report_text


def _write_tiny_png(path: Path) -> None:
    path.write_bytes(
        base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
        )
    )


def _cell_below_header(ws, header: str):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == header:
                return ws.cell(row=cell.row + 1, column=cell.column)
    return None


def _find_cell(ws, value: str):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    return None


def _find_cell_contains(ws, value: str):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and value in cell.value:
                return cell
    return None
