#!/usr/bin/env python3
"""Excel report generator for 复借G卡 model.

Reads training, evaluation, feature, and input snapshots and generates
a formatted Excel report matching the historical 复借G卡模型文档.xlsx structure.

Usage:
    python3 scripts/05_report.py \
      --eval-dir runs/model_eval \
      --train-dir runs/model_train/main_lgbm \
      --input-dir runs/modeling_input \
      --feature-dir runs/modeling_feature_set \
      --output reports/model_report.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd


def find_repo_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "agent.py").exists():
            return candidate
    raise RuntimeError("Cannot locate repo root from script path")


REPO_ROOT = find_repo_root(Path(__file__).resolve())
sys.path.insert(0, str(REPO_ROOT))


# ── style helpers ─────────────────────────────────────────────────


def _style_header(ws, row: int, ncols: int, bold: bool = True, fill_color: str = "4472C4"):
    """Apply header styling to a row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="微软雅黑", size=11, bold=bold, color="FFFFFF")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _style_data_rows(ws, start_row: int, end_row: int, ncols: int):
    """Apply basic styling to data rows."""
    from openpyxl.styles import Font, Alignment, Border, Side

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border


def _write_title(ws, row: int, title: str, ncols: int):
    """Write a merged title row."""
    from openpyxl.styles import Font, Alignment

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(name="微软雅黑", size=14, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_dataframe(ws, start_row: int, df: pd.DataFrame, formats: dict[str, str] | None = None):
    """Write a DataFrame to worksheet starting at start_row.
    Returns the row after the last written row.
    """
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx).value = col_name

    ncols = len(df.columns)
    _style_header(ws, start_row, ncols)

    # Data
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row_data[col_name]
            if isinstance(val, float) and np.isnan(val):
                val = None
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = round(float(val), 6)
            ws.cell(row=start_row + 1 + row_idx, column=col_idx).value = val

    data_end = start_row + len(df)
    _style_data_rows(ws, start_row + 1, data_end, ncols)

    # Auto-width
    for col_idx in range(1, ncols + 1):
        max_width = 8
        for row_idx in range(start_row, data_end + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_width = max(max_width, min(len(str(val)) * 1.2, 40))
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = max_width + 2

    return data_end + 2


def _write_section(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    """Write a titled section and return the next available row."""
    ncols = max(len(df.columns), 1)
    _write_title(ws, start_row, title, ncols)
    return _write_dataframe(ws, start_row + 1, df)


# ── main ──────────────────────────────────────────────────────────


def main() -> None:
    project_dir = Path(__file__).resolve().parents[1]

    parser = argparse.ArgumentParser(description="Generate 复借G卡 model report Excel")
    parser.add_argument("--eval-dir", required=True, help="Evaluation output directory")
    parser.add_argument("--train-dir", required=True, help="Training output directory")
    parser.add_argument("--input-dir", required=True, help="Input snapshot directory")
    parser.add_argument("--feature-dir", required=True, help="Feature set directory")
    parser.add_argument("--output", required=True, help="Output Excel path")
    args = parser.parse_args()

    eval_dir = project_dir / args.eval_dir if not Path(args.eval_dir).is_absolute() else Path(args.eval_dir)
    train_dir = project_dir / args.train_dir if not Path(args.train_dir).is_absolute() else Path(args.train_dir)
    input_dir = project_dir / args.input_dir if not Path(args.input_dir).is_absolute() else Path(args.input_dir)
    feature_dir = project_dir / args.feature_dir if not Path(args.feature_dir).is_absolute() else Path(args.feature_dir)
    output_path = project_dir / args.output if not Path(args.output).is_absolute() else Path(args.output)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook

    wb = Workbook()

    # ── helper: safe read CSV ────────────────────────────────────
    def _read_csv(path: Path) -> pd.DataFrame | None:
        if not path.exists():
            return None
        return pd.read_csv(path, encoding="utf-8-sig")

    # ── load key metrics ─────────────────────────────────────────
    train_config = None
    config_path = train_dir / "run_config.json"
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as fh:
            train_config = json.load(fh)

    metrics = None
    metrics_path = train_dir / "metrics_train_valid.json"
    if metrics_path.exists():
        with open(metrics_path, "r", encoding="utf-8") as fh:
            metrics = json.load(fh)

    eval_summary = None
    summary_path = eval_dir / "evaluation_summary.json"
    if summary_path.exists():
        with open(summary_path, "r", encoding="utf-8") as fh:
            eval_summary = json.load(fh)

    # Feature list
    feature_count = 0
    feature_list_path = feature_dir / "feature_list.txt"
    if feature_list_path.exists():
        with open(feature_list_path, "r", encoding="utf-8") as fh:
            feature_count = sum(1 for line in fh if line.strip())

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    summary_lines = [
        ["复借G卡模型报告"],
        [f"生成日期：{time.strftime('%Y-%m-%d')}"],
        [""],
        ["模型背景"],
        ["本轮是基于新数据、新特征集合、新时间窗口训练的复借G卡主模型。"],
        ["本轮暂不做分客群建模，分客群结果仅作为效果切片评估。"],
        [""],
        ["样本定义"],
        [f"标签字段：ftr_30d_ord_flag（观察日30天内是否发起）"],
        [f"切分字段：final_flag"],
    ]

    if train_config:
        summary_lines += [
            [""],
            ["训练配置"],
            [f"算法：LightGBM (binary)"],
            [f"训练样本：final_flag in {train_config.get('train_values', [])}"],
            [f"验证样本：final_flag in {train_config.get('valid_values', [])}"],
            [f"候选特征数：{train_config.get('candidate_feature_count', 'N/A')}"],
            [f"实际入模特征数：{train_config.get('actual_feature_count', 'N/A')}"],
            [f"Train 样本数：{train_config.get('train_samples', 'N/A')}"],
            [f"Valid 样本数：{train_config.get('valid_samples', 'N/A')}"],
            [f"Train bad rate：{train_config.get('train_bad_rate', 'N/A')}"],
            [f"Valid bad rate：{train_config.get('valid_bad_rate', 'N/A')}"],
            [f"Best iteration：{train_config.get('best_iteration', 'N/A')}"],
        ]

    if metrics:
        summary_lines += [
            [""],
            ["核心效果"],
            [f"Train AUC：{metrics.get('train_auc', 'N/A'):.4f}" if metrics.get('train_auc') else "Train AUC：N/A"],
            [f"Valid AUC：{metrics.get('valid_auc', 'N/A'):.4f}" if metrics.get('valid_auc') else "Valid AUC：N/A"],
            [f"Train KS：{metrics.get('train_ks', 'N/A'):.4f}" if metrics.get('train_ks') else "Train KS：N/A"],
            [f"Valid KS：{metrics.get('valid_ks', 'N/A'):.4f}" if metrics.get('valid_ks') else "Valid KS：N/A"],
            [f"AUC Gap：{metrics.get('auc_gap', 'N/A'):.4f}" if metrics.get('auc_gap') is not None else "AUC Gap：N/A"],
        ]

    summary_lines += [
        [""],
        ["风险观察说明"],
        ["MOB1/MOB3 人头逾期率、金额逾期率的历史计算逻辑待同事确认。"],
        ["当前仅输出金额逾期率和人头风险率作为风险观察，不等同于历史口径。"],
    ]

    for i, line in enumerate(summary_lines, 1):
        if line:
            ws.cell(row=i, column=1).value = line[0]

    # Style the title
    from openpyxl.styles import Font, Alignment
    ws.cell(row=1, column=1).font = Font(name="微软雅黑", size=16, bold=True)
    ws.column_dimensions["A"].width = 80

    # ── Sheet 2: 模型效果 ────────────────────────────────────────
    ws2 = wb.create_sheet("模型效果")
    row = 1

    overall = _read_csv(eval_dir / "overall_metrics.csv")
    if overall is not None:
        row = _write_section(ws2, row, "整体效果 (Overall Metrics)", overall)

    monthly = _read_csv(eval_dir / "monthly_metrics.csv")
    if monthly is not None:
        row = _write_section(ws2, row, "月度效果 (Monthly Metrics)", monthly)

    benchmark = _read_csv(eval_dir / "benchmark_uplift.csv")
    if benchmark is not None:
        row = _write_section(ws2, row, "历史版本对比 (Benchmark Uplift)", benchmark)

    # ── Sheet 3: 分客群效果 ──────────────────────────────────────
    ws3 = wb.create_sheet("分客群效果")
    row = 1
    segment = _read_csv(eval_dir / "segment_metrics.csv")
    if segment is not None:
        row = _write_section(ws3, row, "客群切片效果 (Segment Metrics)", segment)

    note_row = row + 1
    ws3.cell(row=note_row, column=1).value = "注意：分客群结果仅作为切片评估，不代表分客群训练模型。"
    ws3.cell(row=note_row, column=1).font = Font(name="微软雅黑", size=10, italic=True, color="FF0000")

    # ── Sheet 4: Lift ────────────────────────────────────────────
    ws4 = wb.create_sheet("Lift")
    row = 1

    dec_all = _read_csv(eval_dir / "decile_lift_all.csv")
    if dec_all is not None:
        row = _write_section(ws4, row, "全客群十分位 Lift (Decile Lift - All)", dec_all)

    dec_e2e3 = _read_csv(eval_dir / "decile_lift_e2e3.csv")
    if dec_e2e3 is not None:
        row = _write_section(ws4, row, "老户次新十分位 Lift (Decile Lift - E2+E3)", dec_e2e3)

    dec_b2 = _read_csv(eval_dir / "decile_lift_b2.csv")
    if dec_b2 is not None:
        row = _write_section(ws4, row, "流失户十分位 Lift (Decile Lift - B2)", dec_b2)

    # ── Sheet 5: 意愿资产交叉 ────────────────────────────────────
    ws5 = wb.create_sheet("意愿资产交叉")
    row = 1

    intent_zc = _read_csv(eval_dir / "intent_zc_distribution.csv")
    if intent_zc is not None:
        row = _write_section(ws5, row, "意愿评级 x 资产评级交叉 (Intent x Asset Level)", intent_zc)

    ftr_rate = _read_csv(eval_dir / "intent_zc_ftr_rate.csv")
    if ftr_rate is not None:
        row = _write_section(ws5, row, "30天发起率 - 意愿 x 资产 (FTR Rate)", ftr_rate)

    # ── Sheet 6: 风险观察 ────────────────────────────────────────
    ws6 = wb.create_sheet("风险观察")
    row = 1

    amount_risk = _read_csv(eval_dir / "intent_zc_amount_risk.csv")
    if amount_risk is not None:
        row = _write_section(ws6, row, "金额逾期率 (Amount Overdue Rate)", amount_risk)

    head_risk = _read_csv(eval_dir / "intent_zc_headcount_risk.csv")
    if head_risk is not None:
        row = _write_section(ws6, row, "人头风险率 (Headcount Risk Rate)", head_risk)

    note_row = row + 1
    ws6.cell(row=note_row, column=1).value = (
        "注意：当前风险观察仅基于 prc_amt_xz_30d_3m 和 ovd_amt_xz_30d_3m 字段。"
        "MOB1/MOB3 历史口径待同事确认。"
    )
    ws6.cell(row=note_row, column=1).font = Font(name="微软雅黑", size=10, italic=True, color="FF0000")

    # ── Sheet 7: 稳定性 ──────────────────────────────────────────
    ws7 = wb.create_sheet("稳定性")
    row = 1

    psi = _read_csv(eval_dir / "score_psi_by_month.csv")
    if psi is not None:
        row = _write_section(ws7, row, "模型分月度 PSI (Score PSI by Month)", psi)

    # ── Sheet 8: 重要变量 ────────────────────────────────────────
    ws8 = wb.create_sheet("重要变量")
    row = 1

    importance = _read_csv(train_dir / "feature_importance.csv")
    if importance is not None:
        row = _write_section(ws8, row, "特征重要性 (Feature Importance)", importance)

    drop_detail = _read_csv(train_dir / "feature_drop_detail.csv")
    if drop_detail is not None:
        dropped = drop_detail[drop_detail["drop_reason"] != ""]
        if len(dropped) > 0:
            row = _write_section(ws8, row, "被删除特征 (Dropped Features)", dropped)

    # ── Sheet 9: Source Data ─────────────────────────────────────
    ws9 = wb.create_sheet("Source Data")
    row = 1

    # Append all data CSVs as reference
    all_csvs = [
        ("overall_metrics", eval_dir / "overall_metrics.csv"),
        ("monthly_metrics", eval_dir / "monthly_metrics.csv"),
        ("segment_metrics", eval_dir / "segment_metrics.csv"),
        ("decile_lift_all", eval_dir / "decile_lift_all.csv"),
        ("decile_lift_e2e3", eval_dir / "decile_lift_e2e3.csv"),
        ("decile_lift_b2", eval_dir / "decile_lift_b2.csv"),
        ("intent_zc_distribution", eval_dir / "intent_zc_distribution.csv"),
        ("intent_zc_amount_risk", eval_dir / "intent_zc_amount_risk.csv"),
        ("intent_zc_headcount_risk", eval_dir / "intent_zc_headcount_risk.csv"),
        ("score_psi_by_month", eval_dir / "score_psi_by_month.csv"),
        ("benchmark_uplift", eval_dir / "benchmark_uplift.csv"),
        ("feature_importance", train_dir / "feature_importance.csv"),
        ("feature_drop_detail", train_dir / "feature_drop_detail.csv"),
        ("sample_split_summary", input_dir / "sample_split_summary.csv"),
        ("label_distribution", input_dir / "label_distribution.csv"),
        ("segment_distribution", input_dir / "segment_distribution.csv"),
    ]

    for title, csv_path in all_csvs:
        df = _read_csv(csv_path)
        if df is not None and len(df) > 0:
            row = _write_section(ws9, row, f"[{title}] {csv_path.name}", df)

    # ── Save ──────────────────────────────────────────────────────
    wb.save(str(output_path))
    print(f"[REPORT] Saved to {output_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
